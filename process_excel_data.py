"""
Excel数据标准化处理脚本
将报告数据.xlsx标准化为能被pandas顺利读取的格式
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import io

# 设置标准输出编码为UTF-8
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


def find_header_row(sheet, max_search_rows=20):
    """
    查找表头行
    
    Args:
        sheet: openpyxl工作表对象
        max_search_rows: 最大搜索行数
    
    Returns:
        表头行号（从1开始），如果找不到返回None
    """
    header_keywords = ['月份', '时间', '项目', '分类', '名称', '日期', '序号', '编号', 'NO.', 'Energy', 'Date']
    
    for row_idx in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
        row_values = []
        non_empty_count = 0
        
        # 检查前15列
        for col_idx in range(1, min(16, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = str(cell.value) if cell.value is not None else ""
            row_values.append(value)
            if value.strip():
                non_empty_count += 1
        
        # 检查是否包含关键词
        row_str = ' '.join(row_values)
        has_keyword = any(keyword in row_str for keyword in header_keywords)
        
        # 如果包含关键词且至少有3个非空值，认为是表头
        if has_keyword and non_empty_count >= 3:
            return row_idx
    
    # 如果找不到，返回第一个包含多个非空值的行
    for row_idx in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
        non_empty_count = 0
        for col_idx in range(1, min(16, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                non_empty_count += 1
        if non_empty_count >= 3:
            return row_idx
    
    # 如果还是找不到，返回第一行
    return 1


def expand_merged_cells(sheet):
    """
    展开合并单元格，将值填充到所有合并的单元格位置
    
    Args:
        sheet: openpyxl工作表对象
    """
    merged_ranges = list(sheet.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        # 获取合并单元格左上角的值
        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value
        
        # 先取消合并
        sheet.unmerge_cells(str(merged_range))
        
        # 将值填充到所有合并的单元格位置
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = value


def standardize_sheet(sheet, sheet_name):
    """
    标准化单个工作表
    
    Args:
        sheet: openpyxl工作表对象
        sheet_name: 工作表名称
    
    Returns:
        标准化后的DataFrame，如果失败返回None
    """
    try:
        # 1. 展开合并单元格
        expand_merged_cells(sheet)
        
        # 2. 查找表头行
        header_row = find_header_row(sheet)
        print(f"  {sheet_name}: 表头行={header_row}")
        
        # 3. 读取数据（从表头行开始）
        data_rows = []
        for row_idx in range(header_row, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                # 处理日期类型
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                row_data.append(value)
            data_rows.append(row_data)
        
        if not data_rows:
            print(f"  {sheet_name}: 警告 - 没有数据行")
            return None
        
        # 4. 创建DataFrame
        # 第一行作为表头
        header = data_rows[0]
        data = data_rows[1:] if len(data_rows) > 1 else []
        
        # 清理表头
        cleaned_header = []
        for i, col_name in enumerate(header):
            if col_name is None:
                cleaned_header.append(f'列{i+1}')
            else:
                col_str = str(col_name).strip()
                if not col_str or col_str == 'None' or col_str == 'nan':
                    cleaned_header.append(f'列{i+1}')
                else:
                    cleaned_header.append(col_str)
        
        # 确保列名唯一
        seen = {}
        unique_header = []
        for col_name in cleaned_header:
            if col_name in seen:
                seen[col_name] += 1
                unique_header.append(f"{col_name}_{seen[col_name]}")
            else:
                seen[col_name] = 0
                unique_header.append(col_name)
        
        # 创建DataFrame
        if data:
            df = pd.DataFrame(data, columns=unique_header)
        else:
            # 如果没有数据行，创建一个只有表头的DataFrame
            df = pd.DataFrame(columns=unique_header)
        
        # 5. 数据清洗
        # 删除全空行
        df = df.dropna(how='all')
        
        # 删除全空列
        df = df.dropna(axis=1, how='all')
        
        # 删除数据区域末尾的空行
        while len(df) > 0 and df.iloc[-1].isna().all():
            df = df.iloc[:-1]
        
        # 重置索引
        df = df.reset_index(drop=True)
        
        print(f"  {sheet_name}: 标准化完成 - {df.shape[0]} 行 x {df.shape[1]} 列")
        
        return df
        
    except Exception as e:
        print(f"  {sheet_name}: 标准化失败 - {e}")
        import traceback
        traceback.print_exc()
        return None


def process_excel_file(input_path, output_path):
    """
    处理Excel文件，生成标准化版本
    
    Args:
        input_path: 输入Excel文件路径
        output_path: 输出Excel文件路径
    """
    print(f"\n{'='*80}")
    print(f"开始处理Excel文件: {input_path}")
    print(f"{'='*80}\n")
    
    try:
        # 使用openpyxl读取工作簿
        wb = load_workbook(input_path, data_only=True)
        print(f"工作表列表: {wb.sheetnames}\n")
        
        standardized_data = {}
        
        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            print(f"处理工作表: {sheet_name}")
            sheet = wb[sheet_name]
            df = standardize_sheet(sheet, sheet_name)
            if df is not None:
                standardized_data[sheet_name] = df
            print()
        
        # 保存到新的Excel文件
        print(f"\n{'='*80}")
        print(f"保存标准化数据到: {output_path}")
        print(f"{'='*80}\n")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in standardized_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  ✓ {sheet_name}: {df.shape[0]} 行 x {df.shape[1]} 列")
        
        print(f"\n✓ 标准化完成！输出文件: {output_path}")
        print(f"  共处理 {len(standardized_data)} 个工作表")
        
        return True
        
    except Exception as e:
        print(f"\n✗ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def verify_standardized_file(file_path):
    """
    验证标准化后的Excel文件能否被pandas正确读取
    
    Args:
        file_path: Excel文件路径
    """
    print(f"\n{'='*80}")
    print(f"验证标准化文件: {file_path}")
    print(f"{'='*80}\n")
    
    try:
        # 获取工作表列表
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        print(f"工作表列表: {sheet_names}\n")
        
        all_success = True
        
        for sheet_name in sheet_names:
            try:
                # 尝试用pandas读取（使用header=0，因为我们已经标准化了）
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
                
                print(f"✓ {sheet_name}:")
                print(f"  行数: {df.shape[0]}, 列数: {df.shape[1]}")
                print(f"  列名: {list(df.columns)[:10]}")
                if len(df) > 0:
                    print(f"  前3行数据预览:")
                    for idx in range(min(3, len(df))):
                        row_dict = df.iloc[idx].to_dict()
                        # 只显示前5个字段
                        preview = {k: v for i, (k, v) in enumerate(row_dict.items()) if i < 5}
                        print(f"    行{idx+1}: {preview}")
                print()
                
            except Exception as e:
                print(f"✗ {sheet_name}: 读取失败 - {e}\n")
                all_success = False
        
        if all_success:
            print("✓ 所有工作表验证通过！")
        else:
            print("✗ 部分工作表验证失败")
        
        return all_success
        
    except Exception as e:
        print(f"✗ 验证失败: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    input_file = "报告数据.xlsx"
    output_file = "报告数据_标准化.xlsx"
    
    # 处理Excel文件
    success = process_excel_file(input_file, output_file)
    
    if success:
        # 验证标准化后的文件
        verify_standardized_file(output_file)

