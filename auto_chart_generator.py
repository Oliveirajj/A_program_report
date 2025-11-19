"""
Auto-generate charts and tables from the curated JSONL chart logic definitions.

This module reads `chart_generation_logic.jsonl`, loads the standardized Excel
data (`报告数据_标准化.xlsx`), and produces the requested visualizations and tables
under `charts/generated`.

Only ASCII characters are used in source where possible; Chinese literals are
encoded via Unicode escape sequences to keep the file encoding consistent.
"""

from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib
import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.ticker import StrMethodFormatter
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Configure matplotlib with fonts that support Chinese labels (SimHei is common)
matplotlib.rcParams["font.family"] = "sans-serif"
matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS", "DejaVu Sans"]
FONT_CANDIDATES = [
    ("Microsoft YaHei", Path("C:/Windows/Fonts/msyh.ttc")),
    ("Microsoft YaHei", Path("C:/Windows/Fonts/msyh.ttf")),
    ("Arial Unicode MS", Path("C:/Windows/Fonts/arialuni.ttf")),
]
for family, font_path in FONT_CANDIDATES:
    try:
        if font_path.exists():
            font_manager.fontManager.addfont(str(font_path))
    except Exception:
        continue
matplotlib.rcParams["axes.unicode_minus"] = False


ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_LOGIC_PATH = ROOT_DIR / "chart_generation_logic.jsonl"
DEFAULT_EXCEL_PATH = ROOT_DIR / "\u62a5\u544a\u6570\u636e_\u6807\u51c6\u5316.xlsx"
OUTPUT_DIR = ROOT_DIR / "charts" / "generated"
SUBSCRIPT_TWO = "\u2082"
SUPERSCRIPT_TWO = "\u00b2"
COOLING_AIR_SOURCE = "\u98ce\u51b7\u70ed\u6cf5+\u5faa\u73af\u6cf5"
COOLING_MULTI = "\u591a\u8054\u673a\u7a7a\u8c03"
HOT_PUMP_SUMMARY = "\u70ed\u6cf5\u7a7a\u8c03\u5408\u8ba1"
VRF_SUMMARY = "\u591a\u8054\u673a\u7a7a\u8c03\u5408\u8ba1"
LIGHTING_SCOPE_PUBLIC = "public"
LIGHTING_SCOPE_TENANT = "tenant"


class StyleGuide:
    font_family = "Microsoft YaHei"
    header_bg = "#CCFFFF"
    stripe_bg = "#F6FAFF"
    figure_bg = "#FFFFFF"
    grid_color = "#D6E2F3"
    border_color = "BFBFBF"
    text_color = "#1F1F1F"
    secondary_text = "#595959"
    title_color = "#1D3A70"
    primary_blue = "#3C6DC3"
    secondary_blue = "#4473C5"
    accent_orange = "#ED7D31"
    green = "#4EC9B0"

    default_font = Font(name=font_family, size=11, color="000000")
    header_font = Font(name=font_family, size=11, color="000000", bold=True)
    title_font = Font(name=font_family, size=14, color="1F1F1F", bold=True)
    note_font = Font(name=font_family, size=9, color="595959", italic=True)

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    @staticmethod
    def infer_alignment(column: str, alignments: Optional[Dict[str, str]], index: int) -> Alignment:
        if alignments and column in alignments:
            direction = alignments[column]
        elif index == 0:
            direction = "left"
        else:
            direction = "right"
        if direction == "left":
            return StyleGuide.left_alignment
        if direction == "center":
            return StyleGuide.center_alignment
        return StyleGuide.right_alignment

    @staticmethod
    def apply_excel_style(
        sheet,
        df: pd.DataFrame,
        *,
        start_row: int,
        header_row: int,
        excel_formats: Optional[Dict[str, str]] = None,
        column_widths: Optional[Dict[str, float]] = None,
        alignments: Optional[Dict[str, str]] = None,
    ) -> None:
        column_names = list(df.columns)
        for col_idx, column in enumerate(column_names, start=1):
            letter = get_column_letter(col_idx)
            width = column_widths.get(column) if column_widths else None
            if width is None:
                width = max(12, min(40, len(str(column)) * 1.4))
            sheet.column_dimensions[letter].width = width
            header_cell = sheet.cell(row=header_row, column=col_idx)
            header_cell.font = StyleGuide.header_font
            header_cell.fill = PatternFill("solid", fgColor=StyleGuide.header_bg.replace("#", ""))
            header_cell.alignment = StyleGuide.center_alignment
            header_cell.border = StyleGuide.thin_border

        for row_offset, (_, row) in enumerate(df.iterrows(), start=0):
            excel_row = start_row + row_offset
            stripe = row_offset % 2 == 1
            for col_idx, column in enumerate(column_names, start=1):
                cell = sheet.cell(row=excel_row, column=col_idx)
                value = row[column]
                cell.font = StyleGuide.default_font
                cell.alignment = StyleGuide.infer_alignment(column, alignments, col_idx - 1)
                cell.border = StyleGuide.thin_border
                if stripe:
                    cell.fill = PatternFill("solid", fgColor=StyleGuide.stripe_bg.replace("#", ""))
                if excel_formats and column in excel_formats:
                    cell.number_format = excel_formats[column]
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    cell.number_format = "#,##0" if float(value).is_integer() else "#,##0.00"
                else:
                    cell.number_format = "@"
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    @staticmethod
    def format_display_value(value: object, fmt: Optional[object]) -> str:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return ""
        if isinstance(value, (np.floating, np.integer)):
            value = float(value)
        if callable(fmt):
            return fmt(value)
        if fmt is None:
            if isinstance(value, float):
                return format(value, ",.2f")
            return str(value)
        fmt_str = str(fmt)
        try:
            if "{" in fmt_str:
                return fmt_str.format(value)
            return format(value, fmt_str)
        except Exception:
            return str(value)

    @staticmethod
    def make_display_frame(df: pd.DataFrame, display_formats: Optional[Dict[str, object]]) -> pd.DataFrame:
        formatted = {}
        for column in df.columns:
            spec = display_formats.get(column) if display_formats else None
            formatted[column] = [
                StyleGuide._normalize_display_text(StyleGuide.format_display_value(val, spec))
                for val in df[column]
            ]
        return pd.DataFrame(formatted)

    @staticmethod
    def _normalize_display_text(value: object) -> object:
        if isinstance(value, str):
            replacements = {
                SUBSCRIPT_TWO: "$_2$",
                SUPERSCRIPT_TWO: "$^2$",
            }
            for target, replacement in replacements.items():
                if target in value:
                    value = value.replace(target, replacement)
        return value

    @staticmethod
    def export_table_to_excel(
        df: pd.DataFrame,
        path: Path,
        *,
        title: Optional[str] = None,
        excel_formats: Optional[Dict[str, str]] = None,
        column_widths: Optional[Dict[str, float]] = None,
        alignments: Optional[Dict[str, str]] = None,
        notes: Optional[List[str]] = None,
    ) -> None:
        excel_formats = excel_formats or {}
        alignments = alignments or {}
        start_row = 2 if title else 1
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1", startrow=start_row - 1)
            sheet = writer.sheets["Sheet1"]
            if title:
                sheet.cell(row=1, column=1).value = title
                sheet.cell(row=1, column=1).font = StyleGuide.title_font
                sheet.cell(row=1, column=1).alignment = StyleGuide.left_alignment
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            header_row = start_row
            data_start = start_row + 1
            StyleGuide.apply_excel_style(
                sheet,
                df,
                start_row=data_start,
                header_row=header_row,
                excel_formats=excel_formats,
                column_widths=column_widths,
                alignments=alignments,
            )
            if notes:
                note_row = data_start + len(df) + 1
                for idx, note in enumerate(notes, start=0):
                    target_row = note_row + idx
                    sheet.cell(row=target_row, column=1).value = f"* {note}"
                    sheet.cell(row=target_row, column=1).font = StyleGuide.note_font
                    sheet.cell(row=target_row, column=1).alignment = StyleGuide.left_alignment
                    sheet.merge_cells(
                        start_row=target_row,
                        start_column=1,
                        end_row=target_row,
                        end_column=len(df.columns),
                    )
            writer.book.properties.title = title or path.stem

    @staticmethod
    def export_table_to_png(
        df: pd.DataFrame,
        path: Path,
        *,
        title: Optional[str] = None,
        display_formats: Optional[Dict[str, object]] = None,
        alignments: Optional[Dict[str, str]] = None,
        notes: Optional[List[str]] = None,
        body_font_size: float = 11,
        header_font_size: float = 12,
        row_height_scale: float = 1.4,
    ) -> None:
        display_df = StyleGuide.make_display_frame(df, display_formats or {})
        col_count = len(display_df.columns)
        row_count = len(display_df)
        fig_width = max(6, col_count * 1.4)
        base_height = 1.2 if title else 0.6
        note_height = 0.25 * len(notes) if notes else 0
        fig_height = base_height + row_count * 0.55 + note_height
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        fig.patch.set_facecolor(StyleGuide.figure_bg)
        ax.axis("off")
        table = ax.table(
            cellText=display_df.values,
            colLabels=display_df.columns.tolist(),
            cellLoc="center",
            loc="center",
            edges="closed",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(body_font_size)
        table.scale(1, row_height_scale)

        for (row, col), cell in table.get_celld().items():
            cell.set_edgecolor(StyleGuide.grid_color)
            if row == 0:
                cell.set_facecolor(StyleGuide.header_bg)
                cell.set_text_props(
                    fontname=StyleGuide.font_family,
                    color=StyleGuide.text_color,
                    weight="bold",
                    fontsize=header_font_size,
                )
            else:
                if row % 2 == 0:
                    cell.set_facecolor(StyleGuide.stripe_bg)
                else:
                    cell.set_facecolor(StyleGuide.figure_bg)
                align = StyleGuide.infer_alignment(display_df.columns[col], alignments, col) if alignments else None
                text_obj = cell.get_text()
                if align and align.horizontal:
                    text_obj.set_ha(align.horizontal)
                elif col == 0:
                    text_obj.set_ha("left")
                else:
                    text_obj.set_ha("right")
                cell.set_text_props(
                    fontname=StyleGuide.font_family,
                    color=StyleGuide.text_color,
                    fontsize=body_font_size,
                )
        if title:
            fig.text(
                0.01,
                0.99,
                title,
                ha="left",
                va="top",
                fontname=StyleGuide.font_family,
                fontsize=16,
                fontweight="bold",
                color=StyleGuide.title_color,
            )
        if notes:
            note_y = 0.05
            for note in notes:
                fig.text(
                    0.01,
                    note_y,
                    f"* {note}",
                    ha="left",
                    va="bottom",
                    fontname=StyleGuide.font_family,
                    fontsize=10,
                    color=StyleGuide.secondary_text,
                )
                note_y -= 0.03
        fig.tight_layout()
        fig.savefig(path, dpi=300, bbox_inches="tight", facecolor=StyleGuide.figure_bg)
        plt.close(fig)

    @staticmethod
    def apply_plot_theme(ax) -> None:
        ax.set_facecolor(StyleGuide.figure_bg)
        ax.grid(True, axis="y", linestyle="--", color=StyleGuide.grid_color, alpha=0.5)
        for spine in ax.spines.values():
            spine.set_color(StyleGuide.grid_color)
        ax.tick_params(axis="both", colors=StyleGuide.secondary_text, labelsize=10)
        ax.title.set_color(StyleGuide.title_color)
        ax.title.set_fontname(StyleGuide.font_family)
        ax.title.set_fontsize(14)
        ax.xaxis.label.set_fontname(StyleGuide.font_family)
        ax.yaxis.label.set_fontname(StyleGuide.font_family)
        ax.xaxis.label.set_color(StyleGuide.secondary_text)
        ax.yaxis.label.set_color(StyleGuide.secondary_text)
        ax.xaxis.label.set_fontsize(11)
        ax.yaxis.label.set_fontsize(11)


PIE_COLOR_PALETTE = [
    "#1F77B4",
    "#FF7F0E",
    "#2CA02C",
    "#D62728",
    "#9467BD",
    "#8C564B",
    "#E377C2",
    "#7F7F7F",
    "#BCBD22",
    "#17BECF",
]


def parse_float(value: object) -> Optional[float]:
    """Convert a cell value to float, handling thousands separators."""
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace(" ", "").strip()
        numeric = pd.to_numeric(cleaned, errors="coerce")
    else:
        numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return float(numeric)


def parse_str(value: object) -> Optional[str]:
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return None


@dataclass
class ChartLogic:
    chart_id: str
    chart_name: str
    chart_type: str
    description: str
    image_file: str
    chapter: str
    logic_steps: List[str]

    @classmethod
    def from_dict(cls, payload: Dict[str, object]) -> "ChartLogic":
        return cls(
            chart_id=str(payload["chart_id"]),
            chart_name=str(payload["chart_name"]),
            chart_type=str(payload["chart_type"]),
            description=str(payload.get("description", "")),
            image_file=str(payload.get("image_file", "")),
            chapter=str(payload.get("chapter", "")),
            logic_steps=[str(step) for step in payload.get("logic_steps", [])],
        )


class DataRepository:
    """
    Wrapper around the standardized Excel workbook that provides typed accessors
    for the worksheet fragments referenced by the chart logic definitions.
    """

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self._water_table: Optional[pd.DataFrame] = None
        self._water_monthly: Optional[pd.DataFrame] = None
        self._water_summary_total: Optional[float] = None
        self._water_reference_values: Dict[str, Optional[float]] = {"general": None, "advanced": None}
        self._sub_electric_raw: Optional[pd.DataFrame] = None
        self._electric_subsections: Optional[pd.DataFrame] = None
        self._cooling_monthly_breakdown: Optional[pd.DataFrame] = None
        self._lighting_breakdown_table: Optional[pd.DataFrame] = None
        self._lighting_breakdown_total: Optional[float] = None
        self._air_conditioning_breakdown_table: Optional[pd.DataFrame] = None
        self._air_conditioning_breakdown_notes: Optional[List[str]] = None
        self._load_workbook()

    def _load_workbook(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel data workbook not found: {self.excel_path}")
        self._excel_file = pd.ExcelFile(self.excel_path)
        self.info_sheet = self._excel_file.parse("Info", header=None)
        self.total_sheet = self._excel_file.parse("\u603b\u80fd\u8017\u53ca\u78b3\u6392", header=0)
        self.monthly_sheet = self._excel_file.parse("\u9010\u6708\u80fd\u8017", header=0)
        self.sub_electric_sheet = self._excel_file.parse("\u5206\u9879\u7cfb\u7edf\u80fd\u8017\uff08\u7535\uff09", header=0)
        self.sub_gas_sheet = self._excel_file.parse("\u5206\u9879\u7cfb\u7edf\u80fd\u8017\uff08\u71c3\u6c14\uff09", header=0)
        self.water_sheet = self._excel_file.parse("\u6c34", header=0)

    # ------------------------------------------------------------------
    # Generic utilities
    # ------------------------------------------------------------------

    @staticmethod
    def _locate_value(df: pd.DataFrame, keyword: str) -> Optional[Tuple[int, int]]:
        """Return (row, col) indices of the first cell that contains keyword."""
        pattern = keyword
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                if isinstance(value, str) and pattern in value:
                    return row_idx, col_idx
        return None

    @staticmethod
    def _to_numeric(series: Iterable[object]) -> List[Optional[float]]:
        return [pd.to_numeric(value, errors="coerce") for value in series]

    # ------------------------------------------------------------------
    # Specific data extractors
    # ------------------------------------------------------------------

    def get_building_area(self) -> Optional[float]:
        loc = self._locate_value(self.info_sheet, "\u5efa\u7b51\u9762\u79ef")
        if not loc:
            return None
        row_idx, _ = loc
        row = self.info_sheet.iloc[row_idx]
        for value in row.tolist()[::-1]:
            numeric = pd.to_numeric(value, errors="coerce")
            if not math.isnan(numeric):
                return float(numeric)
        return None

    def get_total_summary_row(self) -> pd.Series:
        mask = self.total_sheet["\u65f6\u95f4"] == "2024\u5e741\u6708-12\u6708"
        if mask.sum() == 0:
            raise ValueError("Unable to locate 2024 summary row in total sheet")
        row = self.total_sheet.loc[mask].iloc[0]
        return row

    def get_energy_conversion_table(self) -> pd.DataFrame:
        df = self.total_sheet.copy()
        mask = df["\u65f6\u95f4"].isin(["\u7535\u529b", "\u5929\u7136\u6c14"])
        subset = df.loc[mask, ["\u65f6\u95f4", "用电量\nElectricity", "水\nWater", "用电量\nElectricity_1"]]
        subset.columns = ["\u80fd\u6e90\u540d\u79f0", "\u6298\u6807\u51c6\u7164\u7cfb\u6570", "\u6298\u7b49\u6548\u7535\u7cfb\u6570", "\u78b3\u6392\u653e\u56e0\u5b50"]
        return subset.reset_index(drop=True)

    def get_monthly_energy(self, year: int = 2024) -> pd.DataFrame:
        df = self.monthly_sheet.copy()
        year_col = df.columns[0]
        df["__year"] = pd.to_numeric(df[year_col], errors="coerce").ffill()
        df = df[df["__year"] == year]
        df = df[df["\u6708\u4efd"].astype(str).str.contains("\u6708")]
        df = df[df["\u6708\u4efd"] != "\u5408\u8ba1"]
        df["month"] = df["\u6708\u4efd"].str.replace("\u6708", "", regex=False).astype(int)
        df["electricity_kwh"] = pd.to_numeric(
            df["\u7535\u529b"].astype(str).str.replace(",", "", regex=False), errors="coerce"
        )
        df["gas_m3"] = pd.to_numeric(
            df["\u5929\u7136\u6c14"].astype(str).str.replace(",", "", regex=False), errors="coerce"
        )
        df["gas_equiv_kwh"] = df["gas_m3"] * 7.148
        df = df.sort_values("month")
        return df[["month", "electricity_kwh", "gas_m3", "gas_equiv_kwh"]].reset_index(drop=True)

    def get_energy_composition_summary(self) -> Optional[Dict[str, Optional[float]]]:
        df = self.monthly_sheet.copy()
        if df.empty:
            return None
        month_col = df.columns[1]
        mask = df[month_col] == "2024年能源构成占比分析"
        if not mask.any():
            return None
        row = df.loc[mask].iloc[0]
        electricity = parse_float(row.get("电力"))

        gas_equiv: Optional[float] = None
        for candidate in ["燃气", "天然气", "市政蒸汽"]:
            if candidate in df.columns:
                gas_equiv = parse_float(row.get(candidate))
                if gas_equiv not in (None, 0):
                    break
        return {"electricity_kwh": electricity, "gas_equiv_kwh": gas_equiv}

    def get_monthly_gas(self, year: int = 2024) -> pd.DataFrame:
        df = self.sub_gas_sheet.copy()
        df = df.rename(columns={col: str(col) for col in df.columns})
        header_row = df.iloc[0]
        cols = {
            "month": "Date/ Time",
            "volume": "\u71c3\u6c14\u91cf/m3",
            "price": "\u5355\u4ef7",
            "cost": "合计费用",
        }
        for alias, column in cols.items():
            if column not in df.columns:
                raise KeyError(f"Column '{column}' missing in gas sheet")
        data = df[cols.values()].copy()
        data = data.iloc[1:]
        data["month"] = data["Date/ Time"]
        data = data[data["month"].astype(str).str.contains("\u6708")]
        data = data[data["month"] != "\u5408\u8ba1"]
        data["month"] = data["month"].str.replace("\u6708", "", regex=False).astype(int)
        data["volume"] = pd.to_numeric(data["\u71c3\u6c14\u91cf/m3"], errors="coerce")
        data["price"] = pd.to_numeric(data["\u5355\u4ef7"], errors="coerce")
        data["cost"] = pd.to_numeric(data["合计费用"], errors="coerce")
        return data[["month", "volume", "price", "cost"]].reset_index(drop=True)

    def get_monthly_water(self) -> pd.DataFrame:
        if self._water_monthly is None:
            raw = pd.read_excel(self.excel_path, sheet_name="\u6c34", header=None)
            header = raw.iloc[1, :5].astype(str).str.replace("\n", "")
            monthly = raw.iloc[2:14, :5].copy()
            monthly.columns = header
            monthly.rename(
                columns={
                    "\u5185\u5bb9\u6708\u4efd": "month_label",
                    "\u7528\u6c34\u91cf\uff08m3\uff09": "water_m3",
                    "\u8d39\u7528\uff08\u5143\uff09": "cost",
                    "\u5355\u4ef7\uff08\u5143\uff09": "unit_price",
                    "\u5355\u4f4d\u9762\u79ef\u6bcf\u6708\u7528\u6c34\u91cfm3/\uff08\u33a1\uff0f\u6708\uff09": "intensity",
                },
                inplace=True,
            )
            if "month_label" not in monthly.columns:
                monthly.rename(columns={monthly.columns[0]: "month_label"}, inplace=True)
            monthly["month"] = (
                monthly["month_label"].astype(str).str.replace("\u6708", "", regex=False)
            )
            monthly["month"] = pd.to_numeric(monthly["month"], errors="coerce").astype(int)
            monthly["water_m3"] = monthly["water_m3"].apply(parse_float)
            if "cost" in monthly.columns:
                monthly["cost"] = monthly["cost"].apply(parse_float)
            else:
                monthly["cost"] = np.nan
            if "unit_price" in monthly.columns:
                monthly["unit_price"] = monthly["unit_price"].apply(parse_float)
            else:
                monthly["unit_price"] = np.nan
            if "intensity" in monthly.columns:
                monthly["intensity"] = monthly["intensity"].apply(parse_float)
            else:
                monthly["intensity"] = np.nan
            self._water_monthly = monthly[["month", "water_m3", "cost", "unit_price", "intensity"]].reset_index(drop=True)

            summary_row = raw.iloc[14]
            self._water_summary_total = parse_float(summary_row.iloc[1])

            reference_row = raw.iloc[15].tolist() if len(raw) > 15 else []
            numeric_values = []
            for value in reference_row:
                parsed = parse_float(value)
                if parsed is not None and 0 < parsed < 1:
                    numeric_values.append(parsed)
            numeric_values = sorted(set(numeric_values))
            if numeric_values:
                self._water_reference_values["advanced"] = numeric_values[0]
                if len(numeric_values) > 1:
                    self._water_reference_values["general"] = numeric_values[-1]
                else:
                    self._water_reference_values["general"] = numeric_values[0]

        return self._water_monthly.copy()

    def get_water_summary_row(self) -> pd.Series:
        if self._water_monthly is None:
            self.get_monthly_water()
        return {
            "total_water": self._water_summary_total,
            "general": self._water_reference_values.get("general"),
            "advanced": self._water_reference_values.get("advanced"),
        }

    def get_subsection_electric_table(self) -> pd.DataFrame:
        df = self.sub_electric_sheet.copy()
        df.columns = [str(col) for col in df.columns]
        return df

    def get_subsection_electric_matrix(self) -> pd.DataFrame:
        if self._sub_electric_raw is None:
            self._sub_electric_raw = pd.read_excel(
                self.excel_path, sheet_name="\u5206\u9879\u7cfb\u7edf\u80fd\u8017\uff08\u7535\uff09", header=None
            )
        return self._sub_electric_raw.copy()

    def get_electric_subsections(self) -> pd.DataFrame:
        if self._electric_subsections is None:
            matrix = self.get_subsection_electric_matrix()
            records: List[Dict[str, object]] = []
            month_cols = list(range(16, 28))
            for _, row in matrix.iterrows():
                category = row.get(14)
                item = row.get(15)
                if isinstance(category, str):
                    category = category.strip()
                if isinstance(item, str):
                    item = item.strip()
                months = []
                has_value = False
                for col in month_cols:
                    value = parse_float(row.get(col))
                    months.append(value)
                    if value not in (None, 0):
                        has_value = has_value or value is not None
                total_value = parse_float(row.get(28))
                if not has_value and total_value is None:
                    continue
                records.append(
                    {
                        "category": category,
                        "item": item,
                        **{f"{idx:02d}": months[idx - 1] for idx in range(1, 13)},
                        "total": total_value,
                    }
                )
            self._electric_subsections = pd.DataFrame(records)
        return self._electric_subsections.copy()

    def get_cooling_monthly_breakdown(self) -> pd.DataFrame:
        if self._cooling_monthly_breakdown is None:
            subsections = self.get_electric_subsections()
            month_keys = [f"{idx:02d}" for idx in range(1, 13)]

            def sum_category(category_name: str) -> List[float]:
                rows = subsections[subsections["category"] == category_name]
                totals: List[float] = []
                for key in month_keys:
                    values = pd.to_numeric(rows.get(key), errors="coerce") if not rows.empty else pd.Series(dtype=float)
                    if not rows.empty and values.notna().any():
                        totals.append(float(values.fillna(0).sum()))
                    else:
                        totals.append(0.0)
                return totals

            air_series = sum_category(COOLING_AIR_SOURCE)
            multi_series = sum_category(COOLING_MULTI)
            self._cooling_monthly_breakdown = pd.DataFrame(
                {
                    "month": list(range(1, 13)),
                    "air_source_heat_pump": air_series,
                    "multi_connected_air_conditioner": multi_series,
                }
            )
        return self._cooling_monthly_breakdown.copy()

    def get_cooling_totals(self) -> Dict[str, Optional[float]]:
        subsections = self.get_electric_subsections()

        def total_for(category_name: str) -> Optional[float]:
            rows = subsections[subsections["category"] == category_name]
            if rows.empty:
                return None
            values = pd.to_numeric(rows["total"], errors="coerce").dropna()
            if values.empty:
                return None
            return float(values.sum())

        air_detail = total_for(COOLING_AIR_SOURCE)
        multi_detail = total_for(COOLING_MULTI)
        hotpump_summary = total_for(HOT_PUMP_SUMMARY)
        vrf_summary = total_for(VRF_SUMMARY)

        if hotpump_summary is not None or vrf_summary is not None:
            cooling_total = (hotpump_summary or 0.0) + (vrf_summary or 0.0)
        elif air_detail is not None or multi_detail is not None:
            cooling_total = (air_detail or 0.0) + (multi_detail or 0.0)
        else:
            cooling_total = None

        return {
            "air_source_detail": air_detail,
            "multi_detail": multi_detail,
            "hotpump_summary": hotpump_summary,
            "vrf_summary": vrf_summary,
            "cooling_total": cooling_total,
        }

    def get_cooling_subitems(self, category_name: str) -> pd.DataFrame:
        subsections = self.get_electric_subsections()
        rows = subsections[subsections["category"] == category_name]
        if rows.empty:
            return pd.DataFrame(columns=["item", "total"])
        result = rows[["item", "total"]].copy()
        result["item"] = result["item"].astype(str).str.strip()
        result["total"] = pd.to_numeric(result["total"], errors="coerce")
        result = result.dropna(subset=["total"])
        return result.reset_index(drop=True)

    def get_air_conditioning_summary(self) -> Optional[Dict[str, Optional[float]]]:
        """
        Locate the table with headers '空调总耗电kWh' / '热泵空调kWh' / '多联机空调kWh' and
        return the first data row beneath it.
        """
        headers = ["\u7a7a\u8c03\u603b\u8017\u7535kWh", "\u70ed\u6cf5\u7a7a\u8c03kWh", "\u591a\u8054\u673a\u7a7a\u8c03kWh"]
        sheet = self.sub_electric_sheet
        header_idx: Optional[int] = None

        def normalize(value: Optional[str]) -> Optional[str]:
            if not value:
                return None
            return value.replace(" ", "").replace("\n", "").strip()

        for idx, row in sheet.iterrows():
            normalized_cells = [normalize(parse_str(cell)) for cell in row.tolist()]
            if not any(normalized_cells):
                continue
            if all(
                any(cell and normalize(header) in cell for cell in normalized_cells)
                for header in headers
            ):
                header_idx = idx
                break

        if header_idx is None:
            return None

        header_row = sheet.iloc[header_idx].tolist()

        data_row: Optional[pd.Series] = None
        for idx in range(header_idx + 1, len(sheet)):
            candidate = sheet.iloc[idx]
            if candidate.dropna().empty:
                continue
            data_row = candidate
            break

        if data_row is None:
            return None

        def find_value(keyword: str) -> Optional[float]:
            normalized_keyword = normalize(keyword)
            if not normalized_keyword:
                return None
            for col_idx, header_value in enumerate(header_row):
                header_text = normalize(parse_str(header_value))
                if header_text and normalized_keyword in header_text:
                    return parse_float(data_row.iloc[col_idx])
            return None

        return {
            "cooling_total": find_value(headers[0]),
            "hotpump": find_value(headers[1]),
            "multi": find_value(headers[2]),
        }

    def get_air_conditioning_breakdown(self) -> Tuple[pd.DataFrame, List[str]]:
        if self._air_conditioning_breakdown_table is not None:
            cached_notes = self._air_conditioning_breakdown_notes or []
            return self._air_conditioning_breakdown_table.copy(), list(cached_notes)

        sheet = self.sub_electric_sheet
        notes: List[str] = []
        header_idx: Optional[int] = None
        category_col: Optional[int] = None
        item_col: Optional[int] = None
        value_col: Optional[int] = None

        for idx, row in sheet.iterrows():
            found_keyword = False
            for col_idx, cell in enumerate(row):
                text = parse_str(cell)
                if not text:
                    continue
                normalized = text.replace(" ", "")
                if "空调用电分项" in normalized:
                    found_keyword = True
                    if category_col is None:
                        category_col = col_idx
                    elif item_col is None:
                        item_col = col_idx
                if value_col is None and ("2024" in normalized and ("耗电" in normalized or "kWh" in normalized)):
                    value_col = col_idx
            if found_keyword and category_col is not None and item_col is not None and value_col is not None:
                header_idx = idx
                break

        records: List[Dict[str, object]] = []
        if header_idx is None or category_col is None or item_col is None or value_col is None:
            notes.append("未定位到“空调用电分项”表头，无法解析分项明细。")
        else:
            blank_streak = 0
            for idx in range(header_idx + 1, len(sheet)):
                row = sheet.iloc[idx]
                marker = parse_str(row.iloc[0]) if len(row) > 0 else None
                category = parse_str(row.iloc[category_col]) if category_col < len(row) else None
                item = parse_str(row.iloc[item_col]) if item_col < len(row) else None
                value = parse_float(row.iloc[value_col]) if value_col < len(row) else None
                if not category and not item and value is None:
                    blank_streak += 1
                    if blank_streak >= 2 and records:
                        break
                    continue
                blank_streak = 0
                is_total = bool(
                    (category and "合计" in category)
                    or (item and "合计" in item)
                )
                records.append(
                    {
                        "marker": marker,
                        "category": category,
                        "item": item,
                        "value": value,
                        "is_total": is_total,
                    }
                )
            if not records:
                notes.append("“空调用电分项”表头存在但未读取到任何有效数据行。")

        breakdown_df = pd.DataFrame(records, columns=["marker", "category", "item", "value", "is_total"])

        def _lookup(keyword: str) -> Optional[float]:
            if breakdown_df.empty:
                return None
            mask = breakdown_df["category"].fillna("").str.contains(keyword)
            if not mask.any():
                mask = breakdown_df["item"].fillna("").str.contains(keyword)
            if not mask.any():
                return None
            value = breakdown_df.loc[mask, "value"].iloc[0]
            return value if value is not None and not pd.isna(value) else None

        if breakdown_df.empty:
            detail_sum: Optional[float] = None
        else:
            detail_series = breakdown_df.loc[~breakdown_df["is_total"], "value"].dropna()
            detail_sum = float(detail_series.sum()) if not detail_series.empty else None

        grand_total = _lookup("空调合计")
        if detail_sum is not None and grand_total is not None and not math.isclose(detail_sum, grand_total, rel_tol=0.0, abs_tol=1.0):
            notes.append(
                f"“空调用电分项”明细求和 {detail_sum:,.2f} kWh 与“空调合计” {grand_total:,.2f} kWh 不一致。"
            )

        for keyword in [HOT_PUMP_SUMMARY, VRF_SUMMARY]:
            if _lookup(keyword) is None:
                notes.append(f"未找到“{keyword}”合计行，后续校验可能不完整。")

        self._air_conditioning_breakdown_table = breakdown_df
        self._air_conditioning_breakdown_notes = notes
        return breakdown_df.copy(), list(notes)

    def get_lighting_breakdown_table(self) -> Tuple[pd.DataFrame, Optional[float]]:
        if self._lighting_breakdown_table is None:
            matrix = self.get_subsection_electric_matrix()
            header_idx: Optional[int] = None
            for idx, row in matrix.iterrows():
                header_cell = parse_str(row.get(1))
                if header_cell and "类别" in header_cell:
                    header_idx = idx
                    break

            records: List[Dict[str, object]] = []
            total_value: Optional[float] = None
            if header_idx is not None:
                for idx in range(header_idx + 1, len(matrix)):
                    row = matrix.iloc[idx]
                    category = parse_str(row.get(1))
                    if not category:
                        break
                    if "合计" in category:
                        total_value = parse_float(row.get(4)) or parse_float(row.get(5))
                        break
                    name = parse_str(row.get(3))
                    if not name:
                        continue
                    zone = parse_str(row.get(2))
                    value_2024 = parse_float(row.get(4))
                    zone_total = parse_float(row.get(5))
                    if value_2024 is None:
                        continue
                    records.append(
                        {
                            "category": category,
                            "zone": zone,
                            "name": name,
                            "value_2024": value_2024,
                            "zone_total": zone_total,
                        }
                    )

            self._lighting_breakdown_table = pd.DataFrame(records)
            self._lighting_breakdown_total = total_value

        return self._lighting_breakdown_table.copy(), self._lighting_breakdown_total

    def get_lighting_overview(self) -> Optional[Dict[str, Optional[float]]]:
        loc = self._locate_value(self.sub_electric_sheet, "\u5efa\u7b51\u603b\u8017\u7535/kWh")
        if not loc:
            return None
        row_idx, _ = loc
        header_row = self.sub_electric_sheet.iloc[row_idx]
        data_row_idx = row_idx + 1
        if data_row_idx >= len(self.sub_electric_sheet):
            return None
        data_row = self.sub_electric_sheet.iloc[data_row_idx]
        header_values = header_row.tolist()
        data_values = data_row.tolist()

        def locate_column(keyword: str) -> Optional[int]:
            for idx, value in enumerate(header_values):
                if isinstance(value, str) and keyword in value:
                    return idx
            return None

        def value_for(keyword: str) -> Optional[float]:
            col_idx = locate_column(keyword)
            if col_idx is None:
                return None
            if col_idx >= len(data_values):
                return None
            return parse_float(data_values[col_idx])

        building_total = value_for("\u5efa\u7b51\u603b\u8017\u7535")
        lighting_total = value_for("\u7167\u660e\u603b\u8017\u7535")
        public_lighting = value_for("\u516c\u533a\u7167\u660e")
        tenant_lighting = value_for("\u79df\u6237\u7167\u660e")
        unit_intensity = value_for("\u5355\u4f4d\u9762\u79ef\u7167\u660e\u8017\u7535\u91cf")
        unit_from_sheet = unit_intensity is not None
        if unit_intensity is None and lighting_total is not None:
            building_area = self.get_building_area()
            if building_area:
                unit_intensity = lighting_total / building_area

        return {
            "building_total": building_total,
            "lighting_total": lighting_total,
            "public_lighting": public_lighting,
            "tenant_lighting": tenant_lighting,
            "unit_intensity": unit_intensity,
            "unit_from_sheet": unit_from_sheet,
        }


class ChartGenerator:
    def __init__(
        self,
        logic_path: Path = DEFAULT_LOGIC_PATH,
        excel_path: Path = DEFAULT_EXCEL_PATH,
        output_dir: Path = OUTPUT_DIR,
    ) -> None:
        self.logic_entries = self._load_logic(logic_path)
        self.data_repo = DataRepository(excel_path)
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._lighting_items_cache: Optional[List[Tuple[str, float]]] = None
        self._lighting_total_cache: Optional[float] = None
        self.handlers: Dict[str, Callable[[ChartLogic], None]] = {
            "CH001": self._generate_ch001,
            "CH002": self._generate_ch002,
            "CH003": self._generate_ch003,
            "CH004": self._generate_ch004,
            "CH005": self._generate_ch005,
            "CH006": self._generate_ch006,
            "CH007": self._generate_electric_breakdown_table,
            "CH008": self._generate_ch007,
            "CH009": self._generate_ch008,
            "CH010": self._generate_ch009,
            "CH011": self._generate_ch010,
            "CH012": self._generate_ch011,
            "CH013": self._generate_ch020,
            "CH013A": self._generate_ch021,
            "CH013B": self._generate_ch022,
            "CH014": self._generate_ch023,
            "CH015": self._generate_ch024,
            "CH016": self._generate_lighting_overview_table,
            "CH016A": self._generate_ch019,
            "CH016B": self._generate_lighting_zone_table,
            "CH017": self._generate_lighting_tenant_pie,
            "CH018": self._generate_lighting_public_pie,
            "CH019": self._generate_ch016,
        }

    def _export_table(
        self,
        entry: ChartLogic,
        df: pd.DataFrame,
        *,
        suffix: str,
        title: Optional[str] = None,
        excel_formats: Optional[Dict[str, str]] = None,
        display_formats: Optional[Dict[str, object]] = None,
        column_widths: Optional[Dict[str, float]] = None,
        alignments: Optional[Dict[str, str]] = None,
        notes: Optional[List[str]] = None,
        display_font_size: Optional[float] = None,
        header_font_size: Optional[float] = None,
        row_height_scale: Optional[float] = None,
    ) -> None:
        sanitized_suffix = suffix or "table"
        if not sanitized_suffix.lower().endswith("table"):
            sanitized_suffix = f"{sanitized_suffix}_table"
        base_name = f"{entry.chart_id}_{sanitized_suffix}"
        excel_path = self.output_dir / f"{base_name}.xlsx"
        png_path = self.output_dir / f"{base_name}.png"
        StyleGuide.export_table_to_excel(
            df,
            excel_path,
            title=title or entry.chart_name,
            excel_formats=excel_formats,
            column_widths=column_widths,
            alignments=alignments,
            notes=notes,
        )
        StyleGuide.export_table_to_png(
            df,
            png_path,
            title=title or entry.chart_name,
            display_formats=display_formats,
            alignments=alignments,
            notes=notes,
            body_font_size=display_font_size or 11,
            header_font_size=header_font_size or 12,
            row_height_scale=row_height_scale or 1.4,
        )

    def _render_pie_chart(
        self,
        entry: ChartLogic,
        *,
        labels: Sequence[str],
        values: Sequence[float],
        legend_title: Optional[str],
        output_suffix: str,
        ratios: Optional[Sequence[Optional[float]]] = None,
        palette: Optional[List[str]] = None,
        figsize: Tuple[float, float] = (6.8, 6.8),
    ) -> None:
        if not labels or not values:
            return
        fig, ax = plt.subplots(figsize=figsize)
        fig.patch.set_facecolor(StyleGuide.figure_bg)
        color_cycle = palette or PIE_COLOR_PALETTE
        wedges, _ = ax.pie(
            values,
            labels=None,
            startangle=90,
            colors=color_cycle[: len(values)],
            wedgeprops={"linewidth": 1, "edgecolor": StyleGuide.figure_bg},
            textprops={"fontname": StyleGuide.font_family, "fontsize": 10, "color": StyleGuide.text_color},
        )
        total_value = sum(values)
        legend_entries: List[str] = []
        for idx, (label, value) in enumerate(zip(labels, values)):
            ratio = None
            if ratios and idx < len(ratios):
                ratio = ratios[idx]
            if ratio is None and total_value:
                ratio = value / total_value
            if ratio is not None:
                legend_entries.append(f"{label} ({ratio * 100:.1f}%, {value:,.0f} kWh)")
            else:
                legend_entries.append(f"{label} ({value:,.0f} kWh)")
        legend = ax.legend(
            wedges,
            legend_entries,
            loc="center left",
            bbox_to_anchor=(1.05, 0.5),
            frameon=False,
            prop={"family": StyleGuide.font_family, "size": 11},
        )
        if legend_title:
            legend.set_title(legend_title)
            legend.get_title().set_fontsize(12)
        ax.axis("equal")
        ax.set_title(entry.chart_name, fontname=StyleGuide.font_family, fontsize=14, color=StyleGuide.title_color)
        output_path = self.output_dir / f"{entry.chart_id}_{output_suffix}.png"
        fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor=StyleGuide.figure_bg)
        plt.close(fig)

    def _render_single_bar_chart(
        self,
        entry: ChartLogic,
        *,
        x_positions: Sequence[float],
        values: Sequence[float],
        xtick_labels: Sequence[str],
        bar_label: Optional[str],
        x_label: str,
        y_label: str,
        output_suffix: str,
        color: str,
        legend_loc: str = "upper left",
        annotate: bool = True,
        figsize: Tuple[float, float] = (10, 6),
        ylim: Optional[Tuple[Optional[float], Optional[float]]] = None,
    ) -> None:
        fig, ax = plt.subplots(figsize=figsize)
        fig.patch.set_facecolor(StyleGuide.figure_bg)
        StyleGuide.apply_plot_theme(ax)
        bars = ax.bar(
            x_positions,
            values,
            color=color,
            alpha=0.85,
            label=bar_label,
        )
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label, color=StyleGuide.text_color)
        ax.set_xticks(x_positions)
        ax.set_xticklabels(xtick_labels, fontname=StyleGuide.font_family)
        if ylim:
            ax.set_ylim(bottom=ylim[0], top=ylim[1])
        else:
            ax.set_ylim(bottom=0)
        if bar_label:
            ax.legend(loc=legend_loc, frameon=False)
        if annotate:
            for bar, value in zip(bars, values):
                if pd.isna(value) or value <= 0:
                    continue
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f"{value:,.0f}",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    fontname=StyleGuide.font_family,
                    color=StyleGuide.text_color,
                )
        ax.set_title(entry.chart_name, fontname=StyleGuide.font_family, fontsize=14, color=StyleGuide.title_color)
        fig.tight_layout()
        output_path = self.output_dir / f"{entry.chart_id}_{output_suffix}.png"
        fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor=StyleGuide.figure_bg)
        plt.close(fig)

    def _render_grouped_bar_chart(
        self,
        entry: ChartLogic,
        *,
        x_positions: Sequence[float],
        xtick_labels: Sequence[str],
        series: Sequence[Tuple[str, Sequence[float], str]],
        x_label: str,
        y_label: str,
        output_suffix: str,
        bar_width: float = 0.38,
        legend_loc: str = "upper left",
        annotate: bool = True,
        y_formatter: Optional[StrMethodFormatter] = None,
        ylim: Optional[Tuple[Optional[float], Optional[float]]] = None,
        figsize: Tuple[float, float] = (10, 6),
    ) -> None:
        if not series:
            return
        x = np.array(x_positions, dtype=float)
        fig, ax = plt.subplots(figsize=figsize)
        fig.patch.set_facecolor(StyleGuide.figure_bg)
        StyleGuide.apply_plot_theme(ax)
        total_series = len(series)
        offsets = np.linspace(-(total_series - 1) / 2, (total_series - 1) / 2, total_series) * bar_width
        for offset, (label, values, color) in zip(offsets, series):
            bars = ax.bar(
                x + offset,
                values,
                width=bar_width,
                color=color,
                label=label,
            )
            if annotate:
                for bar, value in zip(bars, values):
                    if pd.isna(value) or value <= 0:
                        continue
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height(),
                        f"{value:,.0f}",
                        ha="center",
                        va="bottom",
                        fontsize=9,
                        fontname=StyleGuide.font_family,
                        color=color,
                    )
        ax.set_xlabel(x_label, fontname=StyleGuide.font_family, color=StyleGuide.text_color)
        ax.set_ylabel(y_label, fontname=StyleGuide.font_family, color=StyleGuide.text_color)
        ax.set_xticks(x)
        ax.set_xticklabels(xtick_labels, fontname=StyleGuide.font_family)
        if y_formatter:
            ax.yaxis.set_major_formatter(y_formatter)
        if ylim:
            ax.set_ylim(bottom=ylim[0], top=ylim[1])
        else:
            ax.set_ylim(bottom=0)
        ax.margins(x=0.015)
        ax.legend(frameon=False, loc=legend_loc)
        ax.set_title(entry.chart_name, fontname=StyleGuide.font_family, fontsize=14, color=StyleGuide.title_color)
        fig.tight_layout()
        output_path = self.output_dir / f"{entry.chart_id}_{output_suffix}.png"
        fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor=StyleGuide.figure_bg)
        plt.close(fig)

    def _render_bar_with_line(
        self,
        entry: ChartLogic,
        *,
        x_positions: Sequence[float],
        xtick_labels: Sequence[str],
        bar_values: Sequence[float],
        bar_label: str,
        bar_color: str,
        x_label: str,
        bar_y_label: str,
        output_suffix: str,
        bar_ylim: Optional[Tuple[Optional[float], Optional[float]]] = None,
        line_values: Optional[Sequence[Optional[float]]] = None,
        line_label: Optional[str] = None,
        line_color: Optional[str] = None,
        line_ylim: Optional[Tuple[Optional[float], Optional[float]]] = None,
        annotate_bars: bool = False,
        figsize: Tuple[float, float] = (10, 6),
    ) -> None:
        fig, ax = plt.subplots(figsize=figsize)
        fig.patch.set_facecolor(StyleGuide.figure_bg)
        StyleGuide.apply_plot_theme(ax)
        bars = ax.bar(
            x_positions,
            bar_values,
            color=bar_color,
            alpha=0.85,
            label=bar_label,
        )
        ax.set_xlabel(x_label)
        ax.set_ylabel(bar_y_label, color=bar_color)
        ax.tick_params(axis="y", labelcolor=bar_color)
        ax.set_xticks(x_positions)
        ax.set_xticklabels(xtick_labels, fontname=StyleGuide.font_family)
        if bar_ylim:
            ax.set_ylim(bottom=bar_ylim[0], top=bar_ylim[1])
        else:
            ax.set_ylim(bottom=0)
        ax2 = None
        if line_values is not None and any(pd.notna(v) for v in line_values):
            ax2 = ax.twinx()
            ax2.set_facecolor("none")
            ax2.grid(False)
            series = pd.Series(line_values)
            filled = series.fillna(method="ffill").fillna(method="bfill")
            line_plot, = ax2.plot(
                x_positions,
                filled,
                color=line_color or StyleGuide.accent_orange,
                marker="o",
                linewidth=2.2,
                markersize=6,
                label=line_label,
            )
            ax2.set_ylabel(line_label or "", color=line_color or StyleGuide.accent_orange)
            ax2.tick_params(axis="y", labelcolor=line_color or StyleGuide.accent_orange)
            if line_ylim:
                ax2.set_ylim(bottom=line_ylim[0], top=line_ylim[1])
            ax2.spines["right"].set_color(StyleGuide.grid_color)
            if line_label:
                ax2.legend(loc="upper right", frameon=False)
        if bar_label:
            ax.legend([bars], [bar_label], loc="upper left", frameon=False)
        if annotate_bars:
            for bar, value in zip(bars, bar_values):
                if pd.isna(value) or value <= 0:
                    continue
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f"{value:,.0f}",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    fontname=StyleGuide.font_family,
                    color=bar_color,
                )
        ax.set_title(entry.chart_name, fontname=StyleGuide.font_family, fontsize=14, color=StyleGuide.title_color)
        fig.tight_layout()
        output_path = self.output_dir / f"{entry.chart_id}_{output_suffix}.png"
        fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor=StyleGuide.figure_bg)
        plt.close(fig)

    def _load_electric_subsections(self) -> pd.DataFrame:
        return self.data_repo.get_electric_subsections()

    def _get_lighting_breakdown_items(self) -> Tuple[List[Tuple[str, float]], Optional[float]]:
        if self._lighting_items_cache is not None:
            return list(self._lighting_items_cache), self._lighting_total_cache

        table, lighting_total = self.data_repo.get_lighting_breakdown_table()
        items: List[Tuple[str, float]] = []
        if not table.empty:
            for _, row in table.iterrows():
                name = str(row["name"]).strip()
                value = parse_float(row.get("value_2024"))
                if value is None or math.isclose(value, 0.0):
                    continue
                items.append((name, float(value)))

        if not lighting_total and items:
            lighting_total = sum(value for _, value in items)

        self._lighting_items_cache = items
        self._lighting_total_cache = lighting_total
        return list(items), lighting_total

    def _aggregate_building_categories(self) -> Dict[str, Dict[str, Optional[float]]]:
        electric_df = self._load_electric_subsections()
        month_keys = [f"{idx:02d}" for idx in range(1, 13)]

        def get_row(name: str) -> Optional[pd.Series]:
            result = electric_df[electric_df["category"] == name]
            if result.empty:
                return None
            return result.iloc[0]

        def aggregate(rows: List[Optional[pd.Series]]) -> Dict[str, Optional[float]]:
            totals: Dict[str, Optional[float]] = {key: 0.0 for key in month_keys}
            totals["total"] = 0.0
            has_data = False
            for row in rows:
                if row is None:
                    continue
                for key in month_keys:
                    value = row.get(key)
                    if value is None or pd.isna(value):
                        continue
                    totals[key] = (totals[key] or 0) + float(value)
                    has_data = True
                total_value = row.get("total")
                if total_value is not None and not pd.isna(total_value):
                    totals["total"] = (totals["total"] or 0) + float(total_value)
                    has_data = True
            if not has_data:
                for key in totals:
                    totals[key] = None
            return totals

        category_map = {
            "\u7167\u660e": ["\u7167\u660e\u5408\u8ba1"],
            "\u7a7a\u8c03": ["\u70ed\u6cfb\u7a7a\u8c03\u5408\u8ba1", "\u591a\u8054\u673a\u7a7a\u8c03\u5408\u8ba1"],
            "\u7535\u68af": ["\u7535\u68af\u5408\u8ba1"],
            "\u5176\u4ed6": [
                "\u6c34\u6cf5\u5408\u8ba1",
                "\u9001\u6392\u98ce\u5408\u8ba1",
                "\u53a8\u623f\u7535\u529b\u5408\u8ba1",
                "\u4fe1\u606f\u3001\u901a\u4fe1\u673a\u623f\u5408\u8ba1",
                "\u5b89\u9632\u5408\u8ba1",
                "\u53d8\u7535\u6240\u5408\u8ba1",
                "\u5145\u7535\u6869\u5408\u8ba1",
                "\u673a\u68b0\u8f66\u5e93\u8bbe\u5907\u5408\u8ba1",
                "\u5176\u4ed6\u5408\u8ba1",
            ],
            "\u5efa\u7b51\u603b\u80fd\u8017": ["\u6284\u8868\u5408\u8ba1"],
            "\u516c\u533a\u5408\u8ba1": ["\u516c\u533a\u5408\u8ba1"],
            "\u79df\u6237\u5408\u8ba1": ["\u79df\u6237\u5408\u8ba1"],
            "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": ["\u4e1a\u4e3b\u5408\u8ba1"],
        }
        aggregated: Dict[str, Dict[str, Optional[float]]] = {}
        for label, keys in category_map.items():
            rows = [get_row(name) for name in keys]
            aggregated[label] = aggregate(rows)
        return aggregated

    def _collect_lighting_scope_items(
        self, target_scope: str, item_whitelist: Optional[List[str]] = None
    ) -> List[Tuple[str, float]]:
        slices: List[Tuple[str, float]] = []
        items, _ = self._get_lighting_breakdown_items()
        whitelist_set = set(item_whitelist) if item_whitelist else None

        tenant_keywords = ["\u7ad6\u4e95", "\u516c\u5171", "\u516c\u533a", "\u5907\u7528"]
        public_keywords = ["\u591c\u666f", "\u5e94\u6025"]

        def matches_scope(name: str) -> bool:
            if target_scope == LIGHTING_SCOPE_TENANT:
                if any(keyword in name for keyword in public_keywords):
                    return False
                return any(keyword in name for keyword in tenant_keywords)
            if target_scope == LIGHTING_SCOPE_PUBLIC:
                return any(keyword in name for keyword in public_keywords)
            return False

        for item_name, value in items:
            if math.isclose(value, 0.0):
                continue
            if whitelist_set is not None and item_name not in whitelist_set:
                continue
            if whitelist_set is None and not matches_scope(item_name):
                continue
            slices.append((item_name, value))

        slices.sort(key=lambda pair: pair[1], reverse=True)
        return slices

    # ------------------------------------------------------------------
    # Loader helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _load_logic(path: Path) -> List[ChartLogic]:
        if not path.exists():
            raise FileNotFoundError(f"Chart logic JSONL not found: {path}")
        entries: List[ChartLogic] = []
        with path.open("r", encoding="utf-8") as fh:
            for line in fh:
                if not line.strip():
                    continue
                payload = json.loads(line)
                entries.append(ChartLogic.from_dict(payload))
        return entries

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate_all(self) -> None:
        for entry in self.logic_entries:
            handler = self.handlers.get(entry.chart_id)
            if handler is None:
                print(f"[WARN] No handler implemented for {entry.chart_id} - {entry.chart_name}")
                continue
            print(f"[INFO] Generating {entry.chart_id} - {entry.chart_name}")
            handler(entry)

    # ------------------------------------------------------------------
    # Chart specific implementations (subset)
    # ------------------------------------------------------------------

    def _generate_ch001(self, entry: ChartLogic) -> None:
        summary = self.data_repo.get_total_summary_row()
        building_area = self.data_repo.get_building_area()

        electricity = parse_float(summary.get("用电量\nElectricity"))
        water = parse_float(summary.get("水\nWater"))
        gas = parse_float(summary.get("燃气\nGas"))
        equivalence = parse_float(summary.get("等效电\nEquivalent Electricity"))
        if equivalence is None and gas is not None:
            equivalence = gas * 7.148
        total_emission = parse_float(summary.get("\u603b\u78b3\u6392\u653e\u91cf"))
        emission_intensity = parse_float(summary.get("\u78b3\u6392\u653e\u5f3a\u5ea6"))
        if emission_intensity is None and total_emission is not None and building_area:
            emission_intensity = total_emission / building_area
        intensity = None
        if building_area and electricity is not None:
            intensity = electricity / building_area

        records = [
            ("\u603b\u7528\u7535\u91cf", electricity, "kWh"),
            ("\u603b\u7528\u6c34\u91cf", water, "m3"),
            ("\u603b\u71c3\u6c14\u91cf", gas, "m3"),
            ("\u603b\u7b49\u6548\u7535", equivalence, "kWh"),
            ("\u603b\u78b3\u6392\u653e\u91cf", total_emission, "t CO2e"),
            ("\u78b3\u6392\u653e\u5f3a\u5ea6", emission_intensity, "t CO2e/m2"),
        ]
        if intensity is not None:
            records.append(("\u5355\u4f4d\u9762\u79ef\u7edf\u80fd", intensity, "kWh/m2"))

        df = pd.DataFrame(records, columns=["指标", "数值", "单位"])
        excel_formats = {"数值": "#,##0.00"}
        display_formats = {"数值": ",.2f"}
        alignments = {"指标": "left", "数值": "right", "单位": "center"}
        notes: List[str] = []
        if building_area:
            notes.append(f"\u5efa\u7b51\u9762\u79ef\uff1a{building_area:,.0f} m2")
        self._export_table(
            entry,
            df,
            suffix="table",
            title=entry.chart_name,
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"指标": 20, "数值": 18, "单位": 12},
            alignments=alignments,
            notes=notes or None,
        )

    def _generate_ch002(self, entry: ChartLogic) -> None:
        df = self.data_repo.get_energy_conversion_table()
        numeric_cols = ["\u6298\u6807\u51c6\u7164\u7cfb\u6570", "\u6298\u7b49\u6548\u7535\u7cfb\u6570", "\u78b3\u6392\u653e\u56e0\u5b50"]
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        excel_formats = {
            "\u6298\u6807\u51c6\u7164\u7cfb\u6570": "0.000",
            "\u6298\u7b49\u6548\u7535\u7cfb\u6570": "0.000",
            "\u78b3\u6392\u653e\u56e0\u5b50": "0.000",
        }
        display_formats = {
            "\u6298\u6807\u51c6\u7164\u7cfb\u6570": ",.3f",
            "\u6298\u7b49\u6548\u7535\u7cfb\u6570": ",.3f",
            "\u78b3\u6392\u653e\u56e0\u5b50": ",.3f",
        }
        alignments = {
            "\u80fd\u6e90\u540d\u79f0": "left",
            "\u6298\u6807\u51c6\u7164\u7cfb\u6570": "right",
            "\u6298\u7b49\u6548\u7535\u7cfb\u6570": "right",
            "\u78b3\u6392\u653e\u56e0\u5b50": "right",
        }
        self._export_table(
            entry,
            df,
            suffix="conversion_table",
            title=entry.chart_name,
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u80fd\u6e90\u540d\u79f0": 18,
                "\u6298\u6807\u51c6\u7164\u7cfb\u6570": 18,
                "\u6298\u7b49\u6548\u7535\u7cfb\u6570": 20,
                "\u78b3\u6392\u653e\u56e0\u5b50": 18,
            },
            alignments=alignments,
        )

    def _generate_ch003(self, entry: ChartLogic) -> None:
        summary = self.data_repo.get_total_summary_row()
        building_area = self.data_repo.get_building_area()

        electricity_kwh = parse_float(summary.get("用电量\nElectricity"))
        gas_m3 = parse_float(summary.get("燃气\nGas"))
        gas_equiv = parse_float(summary.get("等效电\nEquivalent Electricity"))
        if gas_equiv is None and gas_m3 is not None:
            gas_equiv = gas_m3 * 7.148

        electricity_emission = parse_float(summary.get("碳排放量（用电）"))
        if electricity_emission is None and electricity_kwh is not None:
            electricity_emission = electricity_kwh / 1000 * 0.42

        gas_emission = parse_float(summary.get("碳排放量（燃气）"))
        if gas_emission is None and gas_equiv is not None:
            gas_emission = gas_equiv / 1000 * 2.165

        total_emission = parse_float(summary.get("总碳排放量"))
        if total_emission is None and (
            electricity_emission is not None or gas_emission is not None
        ):
            total_emission = (electricity_emission or 0) + (gas_emission or 0)

        gross_ft2 = None
        if building_area and building_area > 0:
            gross_ft2 = building_area / 0.092903
        intensity_kg_ft2: Optional[float] = None
        if total_emission is not None and gross_ft2:
            intensity_kg_ft2 = total_emission * 1000 / gross_ft2

        headers = [
            "时间_",
            "用电量Electricity_",
            "碳排放量（用电）",
            "燃气Gas_",
            "碳排放量（燃气）",
            "总碳排放量",
            "碳排放强度",
        ]
        unit_row = [
            "时间",
            "kWh",
            f"t CO{SUBSCRIPT_TWO}",
            "m³",
            f"t CO{SUBSCRIPT_TWO}",
            f"t CO{SUBSCRIPT_TWO}",
            f"kgCO{SUBSCRIPT_TWO}e/ft{SUPERSCRIPT_TWO}",
        ]
        value_row = [
            "2024年1月-12月",
            electricity_kwh,
            electricity_emission,
            gas_m3,
            gas_emission,
            total_emission,
            intensity_kg_ft2,
        ]
        df = pd.DataFrame([unit_row, value_row], columns=headers)
        excel_formats = {
            "用电量Electricity_": "#,##0",
            "碳排放量（用电）": "#,##0.0000",
            "燃气Gas_": "#,##0.00",
            "碳排放量（燃气）": "#,##0.0000",
            "总碳排放量": "#,##0.0000",
            "碳排放强度": "#,##0.00",
        }
        display_formats = {
            "用电量Electricity_": ",.0f",
            "碳排放量（用电）": ",.4f",
            "燃气Gas_": ",.2f",
            "碳排放量（燃气）": ",.7f",
            "总碳排放量": ",.6f",
            "碳排放强度": ",.2f",
        }
        alignments = {column: "right" for column in headers}
        alignments["时间_"] = "left"
        self._export_table(
            entry,
            df,
            suffix="emission_table",
            title=entry.chart_name,
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "时间_": 16,
                "用电量Electricity_": 20,
                "碳排放量（用电）": 20,
                "燃气Gas_": 16,
                "碳排放量（燃气）": 20,
                "总碳排放量": 20,
                "碳排放强度": 22,
            },
            alignments=alignments,
        )

    def _generate_ch004(self, entry: ChartLogic) -> None:
        df = self.data_repo.get_monthly_energy()
        df = df.dropna(subset=["electricity_kwh", "gas_equiv_kwh"], how="all")
        if df.empty:
            print(f"[WARN] Skipping {entry.chart_id} chart because both series are missing")
            return

        months = df["month"].tolist()
        month_labels = [f"{m}\u6708" for m in months]
        positions = np.arange(len(months))
        series = [
            ("\u7528\u7535\u91cf (kWh)", df["electricity_kwh"].tolist(), StyleGuide.primary_blue),
            ("\u5929\u7136\u6c14\u7b49\u6548\u7535 (kWh)", df["gas_equiv_kwh"].tolist(), StyleGuide.accent_orange),
        ]
        self._render_grouped_bar_chart(
            entry,
            x_positions=positions,
            xtick_labels=month_labels,
            series=series,
            x_label="\u6708\u4efd",
            y_label="kWh",
            output_suffix="grouped_bar",
            y_formatter=StrMethodFormatter("{x:,.0f}"),
        )

    def _generate_ch005(self, entry: ChartLogic) -> None:
        composition = self.data_repo.get_energy_composition_summary() or {}
        electricity = composition.get("electricity_kwh")
        gas_equiv = composition.get("gas_equiv_kwh")
        if gas_equiv is None or gas_equiv <= 0:
            summary = self.data_repo.get_total_summary_row()
            gas_equiv = parse_float(summary.get("等效电\nEquivalent Electricity"))
            if gas_equiv is None:
                gas_m3 = parse_float(summary.get("燃气\nGas"))
                if gas_m3 is not None:
                    gas_equiv = gas_m3 * 7.148
            electricity = electricity or parse_float(summary.get("用电量\nElectricity"))
        pairs = [
            ("\u7535\u529b", electricity),
            ("\u5929\u7136\u6c14", gas_equiv),
        ]
        pairs = [(label, value) for label, value in pairs if not pd.isna(value) and value > 0]
        if not pairs:
            print(f"[WARN] Skipping {entry.chart_id} pie chart because of missing data")
            return
        labels, sizes = zip(*pairs)
        total_size = sum(sizes)
        ratios = [value / total_size if total_size else None for value in sizes]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=sizes,
            ratios=ratios,
            legend_title="\u80fd\u6e90\u7c7b\u578b",
            output_suffix="pie",
            palette=[StyleGuide.primary_blue, StyleGuide.accent_orange],
        )
        total = sum(sizes)
        percent = [value / total if total else 0 for value in sizes]
        df = pd.DataFrame(
            {
                "能源类型": labels,
                "能耗 (kWh)": sizes,
                "占比": percent,
            }
        )
        excel_formats = {"能耗 (kWh)": "#,##0", "占比": "0.0%"}
        display_formats = {"能耗 (kWh)": ",.0f", "占比": ".1%"}
        alignments = {"能源类型": "left", "能耗 (kWh)": "right", "占比": "right"}
        self._export_table(
            entry,
            df,
            suffix="share_table",
            title=f"{entry.chart_name}（数据表）",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"能源类型": 16, "能耗 (kWh)": 18, "占比": 12},
            alignments=alignments,
        )

    def _generate_ch006(self, entry: ChartLogic) -> None:
        df = self.data_repo.get_monthly_energy()
        df = df.dropna(subset=["electricity_kwh"])
        df = df[df["month"].notna()]
        if df.empty:
            print(f"[WARN] {entry.chart_id} 缺少逐月用电量数据，跳过生成")
            return
        df = df.sort_values("month")
        months = df["month"].astype(int).tolist()
        labels = [f"{month}月" for month in months]
        values = df["electricity_kwh"].fillna(0).tolist()
        if not any(values):
            print(f"[WARN] {entry.chart_id} 逐月用电量均为 0，跳过生成")
            return

        self._render_single_bar_chart(
            entry,
            x_positions=months,
            values=values,
            xtick_labels=labels,
            bar_label=None,
            x_label="月份",
            y_label="用电量 (kWh)",
            output_suffix="monthly_electricity",
            color=StyleGuide.primary_blue,
            annotate=False,
        )

        table = pd.DataFrame({"月份": labels, "用电量 (kWh)": values})
        excel_formats = {"用电量 (kWh)": "#,##0"}
        display_formats = {"用电量 (kWh)": ",.0f"}
        alignments = {"月份": "center", "用电量 (kWh)": "right"}
        column_widths = {"月份": 10, "用电量 (kWh)": 18}
        self._export_table(
            entry,
            table,
            suffix="monthly_electricity_table",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths=column_widths,
            alignments=alignments,
        )

    def _generate_electric_breakdown_table(self, entry: ChartLogic) -> None:
        aggregated = self._aggregate_building_categories()
        total_building = aggregated.get("\u5efa\u7b51\u603b\u80fd\u8017", {}).get("total")
        public_total = aggregated.get("\u516c\u533a\u5408\u8ba1", {}).get("total")
        tenant_total = aggregated.get("\u79df\u6237\u5408\u8ba1", {}).get("total")
        owner_total = aggregated.get("\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f", {}).get("total")

        def percent_text(value: Optional[float], denominator: Optional[float]) -> str:
            if value is None or denominator in (None, 0):
                return "--"
            return f"{value / denominator:.1%}"

        table_rows = [
            {
                "\u9879\u76ee": "\u5e74\u8017\u7535 (kWh)",
                "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": total_building,
                "\u516c\u533a": public_total,
                "\u79df\u6237": tenant_total,
                "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": owner_total,
            },
            {
                "\u9879\u76ee": "\u5360\u6bd4",
                "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": "100%",
                "\u516c\u533a": percent_text(public_total, total_building),
                "\u79df\u6237": percent_text(tenant_total, total_building),
                "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": percent_text(owner_total, total_building),
            },
        ]
        df = pd.DataFrame(table_rows)

        excel_formats = {
            "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": "#,##0",
            "\u516c\u533a": "#,##0",
            "\u79df\u6237": "#,##0",
            "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": "#,##0",
        }
        display_formats = {
            "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": ",.0f",
            "\u516c\u533a": ",.0f",
            "\u79df\u6237": ",.0f",
            "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": ",.0f",
        }
        alignments = {
            "\u9879\u76ee": "left",
            "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": "right",
            "\u516c\u533a": "right",
            "\u79df\u6237": "right",
            "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": "right",
        }

        notes: List[str] = []
        if total_building and public_total and tenant_total and owner_total:
            diff = total_building - (public_total + tenant_total + owner_total)
            if abs(diff) > 1:
                notes.append(f"\u63d0\u793a\uff1a\u4e09\u65b9\u603b\u548c\u4e0e\u6284\u8868\u603b\u8017\u7535\u5dee\u503c {diff:,.0f} kWh")

        self._export_table(
            entry,
            df,
            suffix="electric_breakdown",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u9879\u76ee": 16,
                "\u5efa\u7b51\u603b\u8017\u7535\uff08\u6284\u8868\uff09": 22,
                "\u516c\u533a": 16,
                "\u79df\u6237": 16,
                "\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f": 20,
            },
            alignments=alignments,
            notes=notes or None,
        )

    def _generate_ch007(self, entry: ChartLogic) -> None:
        matrix = self.data_repo.get_subsection_electric_matrix()
        header_row = matrix.iloc[10].fillna("")
        value_row = matrix.iloc[11].fillna("")
        ratio_row = matrix.iloc[12].fillna("")
        categories: List[Dict[str, object]] = []
        total_value = parse_float(value_row.iloc[1])
        total_ratio = parse_float(ratio_row.iloc[1]) or 1.0
        categories.append(
            {
                "\u5206\u9879": "\u516c\u533a\u603b\u7528\u7535",
                "\u5e74\u8017\u7535 (kWh)": total_value,
                "\u5360\u516c\u533a\u603b\u7528\u7535": total_ratio or 1.0,
            }
        )
        for idx in range(2, 13):
            name = str(header_row.iloc[idx]).replace("\n", "")
            if not name or name == "nan":
                continue
            value = parse_float(value_row.iloc[idx])
            ratio = parse_float(ratio_row.iloc[idx])
            categories.append(
                {
                    "\u5206\u9879": name,
                    "\u5e74\u8017\u7535 (kWh)": value,
                    "\u5360\u516c\u533a\u603b\u7528\u7535": ratio,
                }
            )
        df = pd.DataFrame(categories)
        excel_formats = {"\u5e74\u8017\u7535 (kWh)": "#,##0", "\u5360\u516c\u533a\u603b\u7528\u7535": "0.0%"}
        display_formats = {"\u5e74\u8017\u7535 (kWh)": ",.0f", "\u5360\u516c\u533a\u603b\u7528\u7535": ".1%"}
        alignments = {
            "\u5206\u9879": "left",
            "\u5e74\u8017\u7535 (kWh)": "right",
            "\u5360\u516c\u533a\u603b\u7528\u7535": "right",
        }
        self._export_table(
            entry,
            df,
            suffix="public_area_breakdown",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u5206\u9879": 28,
                "\u5e74\u8017\u7535 (kWh)": 18,
                "\u5360\u516c\u533a\u603b\u7528\u7535": 18,
            },
            alignments=alignments,
        )

    def _generate_ch008(self, entry: ChartLogic) -> None:
        aggregated = self._aggregate_building_categories()
        total_building = aggregated.get("\u5efa\u7b51\u603b\u80fd\u8017", {}).get("total")
        segments = [
            ("\u516c\u533a", aggregated.get("\u516c\u533a\u5408\u8ba1", {}).get("total")),
            ("\u79df\u6237", aggregated.get("\u79df\u6237\u5408\u8ba1", {}).get("total")),
            ("\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f", aggregated.get("\u4e1a\u4e3b\u4fe1\u606f\u673a\u623f", {}).get("total")),
        ]
        categories = [(label, value) for label, value in segments if value and not math.isclose(value, 0.0)]
        if not categories or not total_building:
            print(f"[WARN] {entry.chart_id} 缺少有效的分项数据，跳过饼图生成")
            return
        labels, values = zip(*categories)
        remainder = total_building - sum(values)
        if abs(remainder) > 1:
            categories.append(("\u5176\u4ed6", remainder))
            labels, values = zip(*categories)
        ratios = [value / total_building if total_building else None for value in values]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title="\u5206\u9879",
            output_suffix="pie",
        )

        df = pd.DataFrame(
            {
                "\u5206\u9879": labels,
                "\u5e74\u8017\u7535 (kWh)": values,
                "\u5360\u603b\u8017\u7535": [val / total_building if total_building else None for val in values],
            }
        )
        excel_formats = {"\u5e74\u8017\u7535 (kWh)": "#,##0", "\u5360\u603b\u8017\u7535": "0.0%"}
        display_formats = {"\u5e74\u8017\u7535 (kWh)": ",.0f", "\u5360\u603b\u8017\u7535": ".1%"}
        alignments = {
            "\u5206\u9879": "left",
            "\u5e74\u8017\u7535 (kWh)": "right",
            "\u5360\u603b\u8017\u7535": "right",
        }
        self._export_table(
            entry,
            df,
            suffix="pie_data",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u5206\u9879": 16,
                "\u5e74\u8017\u7535 (kWh)": 18,
                "\u5360\u603b\u8017\u7535": 14,
            },
            alignments=alignments,
        )

    def _generate_ch009(self, entry: ChartLogic) -> None:
        matrix = self.data_repo.get_subsection_electric_matrix()
        header_row = matrix.iloc[10].fillna("")
        value_row = matrix.iloc[11].fillna("")
        ratio_row = matrix.iloc[12].fillna("")
        labels = []
        values = []
        ratios = []
        for idx in range(2, 12):
            name = str(header_row.iloc[idx]).replace("\n", "")
            if not name or name == "nan":
                continue
            value = parse_float(value_row.iloc[idx])
            ratio = parse_float(ratio_row.iloc[idx])
            if value is None or math.isclose(value, 0.0):
                continue
            labels.append(name)
            values.append(value)
            ratios.append(ratio)
        if not labels:
            print(f"[WARN] {entry.chart_id} 缺少公区分项耗电数据，跳过饼图生成")
            return
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title="\u516c\u533a\u5206\u9879",
            output_suffix="pie",
        )

        df = pd.DataFrame(
            {
                "\u5206\u9879": labels,
                "\u5e74\u8017\u7535 (kWh)": values,
                "\u5360\u6bd4": ratios,
            }
        )
        excel_formats = {"\u5e74\u8017\u7535 (kWh)": "#,##0", "\u5360\u6bd4": "0.0%"}
        display_formats = {"\u5e74\u8017\u7535 (kWh)": ",.0f", "\u5360\u6bd4": ".1%"}
        alignments = {
            "\u5206\u9879": "left",
            "\u5e74\u8017\u7535 (kWh)": "right",
            "\u5360\u6bd4": "right",
        }
        self._export_table(
            entry,
            df,
            suffix="pie_data",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u5206\u9879": 24,
                "\u5e74\u8017\u7535 (kWh)": 18,
                "\u5360\u6bd4": 12,
            },
            alignments=alignments,
        )


    def _generate_lighting_scope_pie(
        self, entry: ChartLogic, target_scope: str, legend_title: str, item_whitelist: Optional[List[str]] = None
    ) -> None:
        slices = self._collect_lighting_scope_items(target_scope, item_whitelist=item_whitelist)
        if not slices:
            print(f"[WARN] {entry.chart_id} \u672a\u627e\u5230{legend_title}\u5206\u9879\u6570\u636e\uff0c\u8df3\u8fc7\u56fe\u8868")
            return
        labels, values = zip(*slices)
        total_value = sum(values)

        ratios = [value / total_value if total_value else None for value in values]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title=legend_title,
            output_suffix="pie",
        )

        table = pd.DataFrame(
            {
                "\u5206\u9879": labels,
                "\u5e74\u8017\u7535 (kWh)": values,
                "\u5360\u5f53\u5e74\u5408\u8ba1": [value / total_value if total_value else None for value in values],
            }
        )
        excel_formats = {"\u5e74\u8017\u7535 (kWh)": "#,##0", "\u5360\u5f53\u5e74\u5408\u8ba1": "0.0%"}
        display_formats = {"\u5e74\u8017\u7535 (kWh)": ",.0f", "\u5360\u5f53\u5e74\u5408\u8ba1": ".1%"}
        alignments = {
            "\u5206\u9879": "left",
            "\u5e74\u8017\u7535 (kWh)": "right",
            "\u5360\u5f53\u5e74\u5408\u8ba1": "right",
        }
        self._export_table(
            entry,
            table,
            suffix="lighting_scope",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u5206\u9879": 24,
                "\u5e74\u8017\u7535 (kWh)": 18,
                "\u5360\u5f53\u5e74\u5408\u8ba1": 18,
            },
            alignments=alignments,
        )

    def _generate_lighting_tenant_pie(self, entry: ChartLogic) -> None:
        self._generate_lighting_scope_pie(entry, LIGHTING_SCOPE_TENANT, "\u79df\u6237\u7167\u660e")

    def _generate_lighting_public_pie(self, entry: ChartLogic) -> None:
        self._generate_lighting_scope_pie(entry, LIGHTING_SCOPE_PUBLIC, "\u516c\u533a\u7167\u660e")

    def _generate_lighting_zone_table(self, entry: ChartLogic) -> None:
        table, lighting_total = self.data_repo.get_lighting_breakdown_table()
        if table.empty:
            print(f"[WARN] {entry.chart_id} 无法提取照明分区明细")
            return
        total_from_sheet = lighting_total is not None
        if lighting_total is None:
            lighting_total = table["value_2024"].fillna(0).sum()

        def normalize_zone(value: object) -> str:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return "未标注分区"
            text = str(value).strip()
            return text or "未标注分区"

        table["category"] = table["category"].fillna("").astype(str).str.strip()
        table["zone"] = table["zone"].apply(normalize_zone)
        table["name"] = table["name"].fillna("").astype(str).str.strip()
        table["value_2024"] = pd.to_numeric(table["value_2024"], errors="coerce")
        table["zone_total"] = pd.to_numeric(table["zone_total"], errors="coerce")
        table = table.dropna(subset=["value_2024"])
        if table.empty:
            print(f"[WARN] {entry.chart_id} 照明分区缺少有效的年耗电数据")
            return

        zone_totals: Dict[str, float] = {}
        fallback_zones: List[str] = []
        for zone, group in table.groupby("zone"):
            sheet_total = group["zone_total"].dropna().iloc[0] if not group["zone_total"].dropna().empty else None
            computed_total = group["value_2024"].sum()
            if sheet_total is None:
                fallback_zones.append(zone)
            zone_totals[zone] = sheet_total if sheet_total is not None else computed_total

        zone_order = sorted(zone_totals.items(), key=lambda item: item[1], reverse=True)
        rows: List[Dict[str, object]] = []
        for zone, zone_total_value in zone_order:
            zone_records = table[table["zone"] == zone].copy()
            zone_records = zone_records.sort_values("value_2024", ascending=False)
            first = True
            for _, record in zone_records.iterrows():
                rows.append(
                    {
                        "\u7c7b\u522b": record["category"] or "-",
                        "\u5206\u533a": zone if first else "",
                        "\u540d\u79f0": record["name"],
                        "2024\u5e74\u8017\u7535 (kWh)": record["value_2024"],
                        "\u5408\u8ba1 (kWh)": zone_total_value if first else None,
                    }
                )
                first = False

        rows.append(
            {
                "\u7c7b\u522b": "",
                "\u5206\u533a": "",
                "\u540d\u79f0": "\u7167\u660e\u5408\u8ba1",
                "2024\u5e74\u8017\u7535 (kWh)": lighting_total,
                "\u5408\u8ba1 (kWh)": lighting_total,
            }
        )

        df = pd.DataFrame(rows)
        excel_formats = {"2024\u5e74\u8017\u7535 (kWh)": "#,##0", "\u5408\u8ba1 (kWh)": "#,##0"}
        display_formats = {"2024\u5e74\u8017\u7535 (kWh)": ",.0f", "\u5408\u8ba1 (kWh)": ",.0f"}
        alignments = {
            "\u7c7b\u522b": "left",
            "\u5206\u533a": "left",
            "\u540d\u79f0": "left",
            "2024\u5e74\u8017\u7535 (kWh)": "right",
            "\u5408\u8ba1 (kWh)": "right",
        }
        notes: List[str] = []
        if fallback_zones:
            notes.append(f"以下分区缺少“合计”值，已按回路求和：{', '.join(sorted(set(fallback_zones)))}")
        if not total_from_sheet:
            notes.append("照明合计按分区求和得出。")

        self._export_table(
            entry,
            df,
            suffix="lighting_zone",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u7c7b\u522b": 12,
                "\u5206\u533a": 16,
                "\u540d\u79f0": 26,
                "2024\u5e74\u8017\u7535 (kWh)": 20,
                "\u5408\u8ba1 (kWh)": 16,
            },
            alignments=alignments,
            notes=notes or None,
            display_font_size=9.5,
            header_font_size=10.5,
            row_height_scale=1.5,
        )

    def _generate_ch016(self, entry: ChartLogic) -> None:
        subsections = self._load_electric_subsections()
        elevator_rows = subsections[subsections["category"] == "\u7535\u68af"]
        elevator_total_row = subsections[subsections["category"] == "\u7535\u68af\u5408\u8ba1"]
        aggregated = self._aggregate_building_categories()
        building_total = aggregated.get("\u5efa\u7b51\u603b\u80fd\u8017", {}).get("total")
        elevator_total = elevator_total_row.iloc[0]["total"] if not elevator_total_row.empty else None

        labels = []
        values = []
        ratios: List[Optional[float]] = []
        table_rows: List[Dict[str, object]] = []
        for _, row in elevator_rows.iterrows():
            item = str(row["item"])
            total = row.get("total")
            if total is None or pd.isna(total) or math.isclose(total, 0.0):
                continue
            share_system = total / elevator_total if elevator_total else None
            share_building = total / building_total if building_total else None
            table_rows.append(
                {
                    "\u7535\u68af\u56de\u8def": item,
                    "\u5e74\u8017\u7535 (kWh)": total,
                    "\u5360\u7535\u68af\u603b\u91cf": share_system,
                    "\u5360\u5efa\u7b51\u603b\u80fd\u8017": share_building,
                }
            )
            # Filter items with <0.2% of elevator total for pie chart
            if share_system is not None and share_system >= 0.002:
                labels.append(item)
                values.append(total)
                ratios.append(share_system)

        if not labels:
            print(f"[WARN] {entry.chart_id} 未找到电梯分项数据")
            return

        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title="\u7535\u68af\u56de\u8def",
            output_suffix="pie",
        )

        if elevator_total:
            table_rows.append(
                {
                    "\u7535\u68af\u56de\u8def": "\u7535\u68af\u603b\u8ba1",
                    "\u5e74\u8017\u7535 (kWh)": elevator_total,
                    "\u5360\u7535\u68af\u603b\u91cf": 1.0,
                    "\u5360\u5efa\u7b51\u603b\u80fd\u8017": (
                        elevator_total / building_total if building_total else None
                    ),
                }
            )
        df = pd.DataFrame(table_rows)
        excel_formats = {
            "\u5e74\u8017\u7535 (kWh)": "#,##0",
            "\u5360\u7535\u68af\u603b\u91cf": "0.0%",
            "\u5360\u5efa\u7b51\u603b\u80fd\u8017": "0.0%",
        }
        display_formats = {
            "\u5e74\u8017\u7535 (kWh)": ",.0f",
            "\u5360\u7535\u68af\u603b\u91cf": ".1%",
            "\u5360\u5efa\u7b51\u603b\u80fd\u8017": ".1%",
        }
        alignments = {
            "\u7535\u68af\u56de\u8def": "left",
            "\u5e74\u8017\u7535 (kWh)": "right",
            "\u5360\u7535\u68af\u603b\u91cf": "right",
            "\u5360\u5efa\u7b51\u603b\u80fd\u8017": "right",
        }
        notes = []
        if elevator_total:
            notes.append(f"\u7535\u68af\u603b\u8017\u7535\uff1a{elevator_total:,.0f} kWh")
        if building_total:
            notes.append(f"\u5efa\u7b51\u603b\u80fd\u8017\uff1a{building_total:,.0f} kWh")
        self._export_table(
            entry,
            df,
            suffix="elevator_breakdown",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u7535\u68af\u56de\u8def": 24,
                "\u5e74\u8017\u7535 (kWh)": 18,
                "\u5360\u7535\u68af\u603b\u91cf": 14,
                "\u5360\u5efa\u7b51\u603b\u80fd\u8017": 18,
            },
            alignments=alignments,
            notes=notes or None,
        )

    def _generate_ch010(self, entry: ChartLogic) -> None:
        df = self.data_repo.get_monthly_gas()
        months = df["month"].tolist()
        month_labels = [f"{m}\u6708" for m in months]
        self._render_bar_with_line(
            entry,
            x_positions=months,
            xtick_labels=month_labels,
            bar_values=df["volume"].tolist(),
            bar_label="\u71c3\u6c14\u91cf (m3)",
            bar_color=StyleGuide.primary_blue,
            x_label="\u6708\u4efd",
            bar_y_label="\u71c3\u6c14\u91cf (m3)",
            output_suffix="gas_line",
            bar_ylim=(0, 3000),
            line_values=df["cost"].tolist(),
            line_label="\u5408\u8ba1\u8d39\u7528 (\u5143)",
            line_color=StyleGuide.accent_orange,
            line_ylim=(0, 14000),
        )
        table = df.copy()
        table.rename(
            columns={
                "month": "\u6708\u4efd",
                "volume": "\u71c3\u6c14\u91cf (m3)",
                "price": "\u5355\u4ef7 (\u5143/m3)",
                "cost": "\u8d39\u7528 (\u5143)",
            },
            inplace=True,
        )
        table["\u6708\u4efd"] = table["\u6708\u4efd"].astype(int)
        excel_formats = {
            "\u71c3\u6c14\u91cf (m3)": "#,##0",
            "\u5355\u4ef7 (\u5143/m3)": "#,##0.00",
            "\u8d39\u7528 (\u5143)": "#,##0",
        }
        display_formats = {
            "\u71c3\u6c14\u91cf (m3)": ",.0f",
            "\u5355\u4ef7 (\u5143/m3)": ",.2f",
            "\u8d39\u7528 (\u5143)": ",.0f",
        }
        alignments = {
            "\u6708\u4efd": "center",
            "\u71c3\u6c14\u91cf (m3)": "right",
            "\u5355\u4ef7 (\u5143/m3)": "right",
            "\u8d39\u7528 (\u5143)": "right",
        }
        self._export_table(
            entry,
            table,
            suffix="gas_table",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={
                "\u6708\u4efd": 10,
                "\u71c3\u6c14\u91cf (m3)": 16,
                "\u5355\u4ef7 (\u5143/m3)": 16,
                "\u8d39\u7528 (\u5143)": 16,
            },
            alignments=alignments,
        )

    def _generate_lighting_overview_table(self, entry: ChartLogic) -> None:
        overview = self.data_repo.get_lighting_overview()
        if not overview:
            print(f"[WARN] {entry.chart_id} 无法从数据源提取照明汇总信息")
            return
        column_map = [
            ("\u5efa\u7b51\u603b\u8017\u7535\n(kWh)", overview.get("building_total")),
            ("\u7167\u660e\u603b\u8017\u7535\n(kWh)", overview.get("lighting_total")),
            ("\u516c\u533a\u7167\u660e\n(kWh)", overview.get("public_lighting")),
            ("\u79df\u6237\u7167\u660e\n(kWh)", overview.get("tenant_lighting")),
            ("\u5355\u4f4d\u9762\u79ef\u7167\u660e\u8017\u7535\u91cf\n(kWh/m\u00b2)", overview.get("unit_intensity")),
        ]
        df = pd.DataFrame([{label: value for label, value in column_map}])
        excel_formats = {
            "\u5efa\u7b51\u603b\u8017\u7535\n(kWh)": "#,##0",
            "\u7167\u660e\u603b\u8017\u7535\n(kWh)": "#,##0",
            "\u516c\u533a\u7167\u660e\n(kWh)": "#,##0",
            "\u79df\u6237\u7167\u660e\n(kWh)": "#,##0",
            "\u5355\u4f4d\u9762\u79ef\u7167\u660e\u8017\u7535\u91cf\n(kWh/m\u00b2)": "#,##0.00",
        }
        display_formats = {
            "\u5efa\u7b51\u603b\u8017\u7535\n(kWh)": ",.0f",
            "\u7167\u660e\u603b\u8017\u7535\n(kWh)": ",.0f",
            "\u516c\u533a\u7167\u660e\n(kWh)": ",.0f",
            "\u79df\u6237\u7167\u660e\n(kWh)": ",.0f",
            "\u5355\u4f4d\u9762\u79ef\u7167\u660e\u8017\u7535\u91cf\n(kWh/m\u00b2)": ",.2f",
        }
        alignments = {column: "right" for column in df.columns}
        notes: List[str] = []
        lighting_total = overview.get("lighting_total") or 0.0
        public_lighting = overview.get("public_lighting") or 0.0
        tenant_lighting = overview.get("tenant_lighting") or 0.0
        diff = lighting_total - (public_lighting + tenant_lighting)
        if lighting_total and abs(diff) > 1.0:
            notes.append("温馨提示：公区+租户照明与照明总耗电存在 {:.0f} kWh 的差值".format(diff))
        if overview.get("unit_intensity") and not overview.get("unit_from_sheet"):
            notes.append("单位面积照明耗电量按“照明总耗电 ÷ 建筑面积”折算。")
        self._export_table(
            entry,
            df,
            suffix="lighting_overview",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={column: 22 for column in df.columns},
            alignments=alignments,
            notes=notes or None,
            display_font_size=9,
            header_font_size=10,
            row_height_scale=1.6,
        )

    def _generate_ch019(self, entry: ChartLogic) -> None:
        overview = self.data_repo.get_lighting_overview()
        if not overview:
            print(f"[WARN] {entry.chart_id} 无法提取照明分项数据")
            return
        lighting_total = overview.get("lighting_total")
        public_value = overview.get("public_lighting")
        tenant_value = overview.get("tenant_lighting")
        if public_value is None and tenant_value is not None and lighting_total is not None:
            public_value = max(lighting_total - tenant_value, 0.0)
        if tenant_value is None and public_value is not None and lighting_total is not None:
            tenant_value = max(lighting_total - public_value, 0.0)
        segments = [
            ("\u516c\u533a\u7167\u660e", public_value),
            ("\u79df\u6237\u7167\u660e", tenant_value),
        ]
        categories = [(label, value) for label, value in segments if value is not None and value > 0]
        if not categories:
            print(f"[WARN] {entry.chart_id} 照明分项数值为空，跳过饼图")
            return
        labels, values = zip(*categories)
        total = sum(values)
        ratios = [value / total if total else None for value in values]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title="\u7167\u660e\u5206\u9879",
            output_suffix="lighting_split",
        )

        table = pd.DataFrame(
            {
                "照明分项": labels,
                "年耗电 (kWh)": values,
                "占照明总耗电": [value / total if total else None for value in values],
            }
        )
        excel_formats = {"年耗电 (kWh)": "#,##0", "占照明总耗电": "0.0%"}
        display_formats = {"年耗电 (kWh)": ",.0f", "占照明总耗电": ".1%"}
        alignments = {"照明分项": "left", "年耗电 (kWh)": "right", "占照明总耗电": "right"}
        self._export_table(
            entry,
            table,
            suffix="lighting_split",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"照明分项": 16, "年耗电 (kWh)": 18, "占照明总耗电": 18},
            alignments=alignments,
        )

    def _generate_ch020(self, entry: ChartLogic) -> None:
        monthly = self.data_repo.get_cooling_monthly_breakdown()
        if monthly.empty:
            print(f"[WARN] {entry.chart_id} 缺少空调用电月度数据")
            return
        monthly = monthly.sort_values("month")
        monthly["label"] = monthly["month"].apply(lambda m: f"{int(m)}\u6708")
        air_values = monthly["air_source_heat_pump"].fillna(0).tolist()
        vrf_values = monthly["multi_connected_air_conditioner"].fillna(0).tolist()
        positions = np.arange(len(monthly))
        series = [
            ("\u98ce\u51b7\u70ed\u6cf5 (kWh)", air_values, StyleGuide.primary_blue),
            ("\u591a\u8054\u673a\u7a7a\u8c03 (kWh)", vrf_values, StyleGuide.accent_orange),
        ]
        self._render_grouped_bar_chart(
            entry,
            x_positions=positions,
            xtick_labels=monthly["label"].tolist(),
            series=series,
            x_label="\u6708\u4efd",
            y_label="\u7a7a\u8c03\u7528\u7535\u91cf (kWh)",
            output_suffix="cooling_monthly",
            y_formatter=StrMethodFormatter("{x:,.0f}"),
            ylim=(0, 100000),
        )

        table = pd.DataFrame(
            {
                "月份": monthly["label"],
                "风冷热泵 (kWh)": air_values,
                "多联机空调 (kWh)": vrf_values,
            }
        )
        excel_formats = {"风冷热泵 (kWh)": "#,##0", "多联机空调 (kWh)": "#,##0"}
        display_formats = {"风冷热泵 (kWh)": ",.0f", "多联机空调 (kWh)": ",.0f"}
        alignments = {"月份": "center", "风冷热泵 (kWh)": "right", "多联机空调 (kWh)": "right"}
        self._export_table(
            entry,
            table,
            suffix="cooling_monthly",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"月份": 10, "风冷热泵 (kWh)": 20, "多联机空调 (kWh)": 20},
            alignments=alignments,
        )

    def _generate_ch021(self, entry: ChartLogic) -> None:
        totals = self.data_repo.get_cooling_totals() or {}
        hotpump = totals.get("hotpump_summary") or totals.get("air_source_detail")
        vrf = totals.get("vrf_summary") or totals.get("multi_detail")
        segments = [
            ("\u70ed\u6cf5\u7a7a\u8c03", hotpump),
            ("\u591a\u8054\u673a\u7a7a\u8c03", vrf),
        ]
        categories = [(label, value) for label, value in segments if value and not math.isclose(value, 0.0)]
        if not categories:
            print(f"[WARN] {entry.chart_id} 空调用电汇总数据缺失，跳过饼图生成")
            return
        labels, values = zip(*categories)
        total_value = sum(values)
        ratios = [value / total_value if total_value else None for value in values]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title="\u7a7a\u8c03\u7cfb\u7edf",
            output_suffix="cooling_split",
        )

        table = pd.DataFrame(
            {
                "空调系统": labels,
                "年耗电 (kWh)": values,
                "占空调总耗电": [value / total_value if total_value else None for value in values],
            }
        )
        excel_formats = {"年耗电 (kWh)": "#,##0", "占空调总耗电": "0.0%"}
        display_formats = {"年耗电 (kWh)": ",.0f", "占空调总耗电": ".1%"}
        alignments = {"空调系统": "left", "年耗电 (kWh)": "right", "占空调总耗电": "right"}
        self._export_table(
            entry,
            table,
            suffix="cooling_split",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"空调系统": 14, "年耗电 (kWh)": 18, "占空调总耗电": 18},
            alignments=alignments,
        )

    def _generate_ch022(self, entry: ChartLogic) -> None:
        summary = self.data_repo.get_air_conditioning_summary()
        breakdown_df, repo_notes = self.data_repo.get_air_conditioning_breakdown()
        notes: List[str] = list(repo_notes)

        total = summary.get("cooling_total") if summary else None
        hotpump = summary.get("hotpump") if summary else None
        multi = summary.get("multi") if summary else None

        def lookup(keyword: str) -> Optional[float]:
            if breakdown_df.empty:
                return None
            mask = breakdown_df["category"].fillna("").str.contains(keyword)
            if not mask.any():
                mask = breakdown_df["item"].fillna("").str.contains(keyword)
            if not mask.any():
                return None
            value = breakdown_df.loc[mask, "value"].iloc[0]
            return value if value is not None and not pd.isna(value) else None

        def reconcile(current: Optional[float], keyword: str, label: str) -> Tuple[Optional[float], Optional[float]]:
            breakdown_value = lookup(keyword)
            if current is None and breakdown_value is not None:
                notes.append(f"表头缺少“{label}”，已使用“{keyword}”行的 {breakdown_value:,.2f} kWh 回填。")
                current = breakdown_value
            elif current is not None and breakdown_value is not None and not math.isclose(current, breakdown_value, rel_tol=0.0, abs_tol=1.0):
                notes.append(f"“{label}” {current:,.2f} kWh 与“{keyword}” {breakdown_value:,.2f} kWh 不一致。")
            return current, breakdown_value

        total, breakdown_total = reconcile(total, "空调合计", "空调总耗电kWh")
        hotpump, _ = reconcile(hotpump, HOT_PUMP_SUMMARY, "热泵空调kWh")
        multi, _ = reconcile(multi, VRF_SUMMARY, "多联机空调kWh")

        if total is None and hotpump is not None and multi is not None:
            total = hotpump + multi
            notes.append("空调总耗电缺失，已使用热泵与多联机合计求和。")

        columns = ["空调总耗电kWh", "热泵空调kWh", "多联机空调kWh"]
        rows: List[Dict[str, Optional[object]]] = [
            {
                "空调总耗电kWh": total,
                "热泵空调kWh": hotpump,
                "多联机空调kWh": multi,
            }
        ]
        if not breakdown_df.empty:
            rows.append({column: None for column in columns})
            rows.append(
                {
                    "空调总耗电kWh": "空调用电分项",
                    "热泵空调kWh": "系统/设备",
                    "多联机空调kWh": "2024年总耗电kWh",
                }
            )
            for _, record in breakdown_df.iterrows():
                rows.append(
                    {
                        "空调总耗电kWh": record.get("category"),
                        "热泵空调kWh": record.get("item"),
                        "多联机空调kWh": record.get("value"),
                    }
                )

        table = pd.DataFrame(rows, columns=columns)
        excel_formats = {
            "空调总耗电kWh": "#,##0.0",
            "热泵空调kWh": "#,##0.0",
            "多联机空调kWh": "#,##0.0",
        }
        display_formats = {
            "空调总耗电kWh": ",.1f",
            "热泵空调kWh": ",.1f",
            "多联机空调kWh": ",.1f",
        }
        alignments = {
            "空调总耗电kWh": "left",
            "热泵空调kWh": "left",
            "多联机空调kWh": "right",
        }
        column_widths = {
            "空调总耗电kWh": 20,
            "热泵空调kWh": 26,
            "多联机空调kWh": 20,
        }

        deduped_notes = list(dict.fromkeys(note for note in notes if note))
        self._export_table(
            entry,
            table,
            suffix="cooling_summary",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths=column_widths,
            alignments=alignments,
            notes=deduped_notes or None,
            header_font_size=12,
            row_height_scale=1.45,
        )

    def _generate_cooling_subsystem_pie(
        self, entry: ChartLogic, category_name: str, system_label: str, suffix: str
    ) -> None:
        items = self.data_repo.get_cooling_subitems(category_name)
        if items.empty:
            print(f"[WARN] {entry.chart_id} 未找到 {system_label} 分项数据，跳过饼图")
            return
        labels = [str(label).strip() or system_label for label in items["item"].tolist()]
        values = [float(value) for value in items["total"].tolist()]
        total_value = sum(values)
        if total_value == 0:
            print(f"[WARN] {entry.chart_id} {system_label} 分项值为 0，跳过饼图")
            return

        ratios = [value / total_value if total_value else None for value in values]
        self._render_pie_chart(
            entry,
            labels=labels,
            values=values,
            ratios=ratios,
            legend_title=system_label,
            output_suffix=suffix,
        )

        table = pd.DataFrame(
            {
                "分项": labels,
                "年耗电 (kWh)": values,
                "占系统耗电": [value / total_value if total_value else None for value in values],
            }
        )
        excel_formats = {"年耗电 (kWh)": "#,##0", "占系统耗电": "0.0%"}
        display_formats = {"年耗电 (kWh)": ",.0f", "占系统耗电": ".1%"}
        alignments = {"分项": "left", "年耗电 (kWh)": "right", "占系统耗电": "right"}
        self._export_table(
            entry,
            table,
            suffix=f"{suffix}_table",
            excel_formats=excel_formats,
            display_formats=display_formats,
            column_widths={"分项": 24, "年耗电 (kWh)": 18, "占系统耗电": 16},
            alignments=alignments,
        )

    def _generate_ch023(self, entry: ChartLogic) -> None:
        self._generate_cooling_subsystem_pie(entry, COOLING_AIR_SOURCE, "\u98ce\u51b7\u70ed\u6cf5+\u5faa\u73af\u6cf5", "air_source_split")

    def _generate_ch024(self, entry: ChartLogic) -> None:
        self._generate_cooling_subsystem_pie(entry, COOLING_MULTI, "\u591a\u8054\u673a\u7a7a\u8c03", "multi_split")

    def _generate_ch011(self, entry: ChartLogic) -> None:
        df = self.data_repo.get_monthly_water()
        months = df["month"].tolist()
        month_labels = [f"{m}\u6708" for m in months]
        self._render_single_bar_chart(
            entry,
            x_positions=months,
            values=df["water_m3"].tolist(),
            xtick_labels=month_labels,
            bar_label="\u7528\u6c34\u91cf (m3)",
            x_label="\u6708\u4efd",
            y_label="\u7528\u6c34\u91cf (m3)",
            output_suffix="water_column",
            color=StyleGuide.green,
            annotate=False,
        )

def main() -> None:
    generator = ChartGenerator()
    generator.generate_all()


if __name__ == "__main__":
    main()

